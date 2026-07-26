"""Inspect official Phase 5G-0 source endpoints without publishing data."""

from __future__ import annotations

import csv
import io
import json
import re
import sys
import urllib.request
import zipfile
from xml.etree import ElementTree

USER_AGENT = "Konsider research source discovery/1.0"


def fetch(url: str) -> tuple[bytes, str]:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=90) as response:
        return response.read(), response.geturl()


def discover_ilostat() -> None:
    url = "https://rplumber.ilo.org/metadata/toc/indicator/?lang=en"
    body, resolved = fetch(url)
    text = body.decode("utf-8-sig")
    try:
        payload = json.loads(text)
        values = payload.get("data", payload)
        if isinstance(values, dict):
            values = values.get("items", values.get("results", []))
        rows = values
    except json.JSONDecodeError:
        rows = csv.DictReader(io.StringIO(text))
    matches = [
        row
        for row in rows
        if "average weekly hours actually worked per employed person"
        in " ".join(str(value) for value in row.values()).lower()
    ]
    print("\nILOSTAT", resolved, len(body), "bytes")
    print(json.dumps(matches, indent=2, ensure_ascii=False))


def discover_who() -> None:
    url = "https://apps.who.int/nha/database/DocumentationCentre/Index/en"
    body, resolved = fetch(url)
    html = body.decode("utf-8", errors="replace")
    print("\nWHO GHED", resolved, len(body), "bytes")
    for term in ("GHED all data", "March 2026", "DocumentationCentre"):
        print(f"\nContext for {term!r}:")
        for match in list(re.finditer(term, html, flags=re.IGNORECASE))[:10]:
            context = re.sub(r"\s+", " ", html[max(0, match.start() - 400) : match.end() + 400])
            print(context.encode("ascii", errors="backslashreplace").decode("ascii"))
    download_url = "https://apps.who.int/nha/database/Home/IndicatorsDownload/en"
    workbook_body, resolved_download = fetch(download_url)
    print("\nWHO workbook", resolved_download, len(workbook_body), "bytes")
    try:
        import openpyxl
    except ImportError:
        return
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_body), read_only=True, data_only=True)
    print("Sheets:", workbook.sheetnames)
    data_sheet = workbook["Data"]
    print("Data sample:")
    for row in data_sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        print(json.dumps(list(row[:20]), ensure_ascii=False, default=str))
    codebook = workbook["Codebook"]
    print(
        "Codebook OOP row:",
        json.dumps(
            [cell.value for cell in next(codebook.iter_rows(min_row=16, max_row=16))],
            ensure_ascii=False,
            default=str,
        ),
    )
    for sheet in workbook.worksheets:
        hits = []
        for row in sheet.iter_rows():
            for cell in row:
                value = str(cell.value or "")
                if "out-of-pocket" in value.lower() or "oop" == value.lower():
                    hits.append((cell.coordinate, value))
            if len(hits) >= 20:
                break
        if hits:
            print(
                f"{sheet.title}: {sheet.max_row} rows x {sheet.max_column} columns; "
                f"hits={json.dumps(hits[:20], ensure_ascii=False)}"
            )


def discover_globalcit() -> None:
    query = "https://globalcit.eu/modes-acquisition-citizenship/"
    body, resolved = fetch(query)
    html = body.decode("utf-8", errors="replace")
    print("\nGLOBALCIT", resolved, len(body), "bytes")
    for match in re.finditer(
        r"""(?P<quote>["'])(?P<url>[^"']*(?:csv|xlsx|ajax|api)[^"']*)(?P=quote)""",
        html,
        flags=re.IGNORECASE,
    ):
        print(match.group("url"))


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return [
        "".join(node.text or "" for node in item.findall(".//a:t", namespace))
        for item in root.findall("a:si", namespace)
    ]


def discover_inform() -> None:
    url = (
        "https://drmkc.jrc.ec.europa.eu/inform-index/Portals/0/InfoRM/2026/"
        "INFORM_Risk_2026_v072.xlsx"
    )
    body, resolved = fetch(url)
    print("\nINFORM", resolved, len(body), "bytes")
    with zipfile.ZipFile(io.BytesIO(body)) as archive:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets = [
            node.attrib["name"] for node in workbook.findall(".//a:sheets/a:sheet", namespace)
        ]
        print("Sheets:", json.dumps(sheets, ensure_ascii=False))
        strings = _xlsx_shared_strings(archive)
        terms = ("river flood", "tropical cyclone", "coastal flood", "drought")
        matches = [
            (index, value)
            for index, value in enumerate(strings)
            if any(term in value.lower() for term in terms)
        ]
        print("Hazard strings:", json.dumps(matches, indent=2, ensure_ascii=False))
    try:
        import openpyxl
    except ImportError:
        return
    workbook = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    for sheet_name in ("INFORM Risk 2026 (a-z)", "Hazard & Exposure"):
        sheet = workbook[sheet_name]
        print(f"\n{sheet_name}: {sheet.max_row} rows x {sheet.max_column} columns")
        for row in sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row), values_only=True):
            print(json.dumps(list(row[:25]), ensure_ascii=False, default=str))


def main() -> int:
    discoveries = {
        "ilostat": discover_ilostat,
        "who": discover_who,
        "globalcit": discover_globalcit,
        "inform": discover_inform,
    }
    requested = sys.argv[1:] or list(discoveries)
    for name in requested:
        try:
            discoveries[name]()
        except Exception as exc:  # noqa: BLE001 - diagnostics should continue
            print(f"\n{name.upper()} ERROR: {type(exc).__name__}: {exc}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
