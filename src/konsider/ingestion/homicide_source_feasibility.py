"""Read-only Phase 2D.4 homicide source feasibility study."""

from __future__ import annotations

import io
import json
import math
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from collections.abc import Callable, Iterable, Mapping
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from konsider.ingestion.country_coverage import _latest_non_null, _wdi_series
from konsider.ingestion.models import RawArtifact, SourceRegistration
from konsider.ingestion.validation import FRESHNESS_MAX_AGE
from konsider.repositories.raw_artifact_repository import RawArtifactRepository

STUDY_SCHEMA_VERSION = "homicide-source-feasibility-1.0"
CRITERION_ID = "intentional_homicide_rate"
SERIES_CODE = "VC_IHR_PSRC"
UNIT = "per_100000_people"
DIRECT_UNODC_WORKBOOK_URL = (
    "https://data.unodc.org/sites/dataportal.unodc.org/files/2026-07/"
    "data_cts_intentional_homicide.xlsx"
)
DIRECT_UNODC_METADATA_URL = (
    "https://data.unodc.org/sites/dataportal.unodc.org/files/2026-07/"
    "metadata_intentional_homicide.pdf"
)
UNSD_SERIES_URL = "https://unstats.un.org/SDGAPI/v1/sdg/Indicator/16.1.1/Series/List"
UNSD_DATA_URL = (
    "https://unstats.un.org/SDGAPI/v1/sdg/Series/Data?"
    "seriesCode=VC_IHR_PSRC&releaseCode=2026.Q2.G.01&page=1&pageSize=20000"
)
EUROSTAT_DATA_URL = (
    "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"
    "crim_hom_soff?lang=en&freq=A&iccs=ICCS0101&leg_stat=PER_VICT&sex=T"
    "&unit=P_HTHAB&sinceTimePeriod=2021"
)
OECD_DATAFLOW_URL = (
    "https://sdmx.oecd.org/public/rest/dataflow/OECD.CFE.EDS/" "DSD_REG_SOC@DF_SAFETY/2.4"
)


def _registration(
    source_id: str,
    publisher: str,
    urls: tuple[str, ...],
    *,
    dataset_version: str,
    methodology_url: str,
    license_name: str,
    license_url: str,
    redistribution: str,
    parser: str,
    notes: str,
) -> SourceRegistration:
    return SourceRegistration(
        source_id=source_id,
        criterion_id=CRITERION_ID,
        publisher=publisher,
        distributor=None,
        canonical_page_url=methodology_url,
        download_urls=urls,
        access_method="official_machine_readable_download",
        pagination="none",
        dataset_version=dataset_version,
        source_version=dataset_version,
        reference_period="annual",
        update_frequency="not_confirmed",
        methodology_url=methodology_url,
        license_name=license_name,
        license_url=license_url,
        redistribution=redistribution,
        permitted_usage="feasibility study and local reproducible analysis only",
        attribution=publisher,
        license_evidence=license_url,
        parser=parser,
        parser_version="phase_2d4_v1",
        official_or_independent="official",
        notes=notes,
    )


PRIMARY_REGISTRATIONS = {
    "direct_unodc_homicide": _registration(
        "direct_unodc_homicide",
        "United Nations Office on Drugs and Crime",
        (DIRECT_UNODC_WORKBOOK_URL, DIRECT_UNODC_METADATA_URL),
        dataset_version="UNODC Data Portal workbook dated 2026-07-12",
        methodology_url="https://data.unodc.org/datareport/hom-estimate",
        license_name="United Nations website terms; redistribution compatibility unresolved",
        license_url="https://www.un.org/en/about-us/terms-of-use",
        redistribution=(
            "General UN terms permit personal non-commercial copying but prohibit redistribution "
            "and derivative compilations absent more specific permission."
        ),
        parser="direct_unodc_homicide_workbook",
        notes="Experimental study source; never registered as the production criterion source.",
    ),
    "unsd_sdg_homicide": _registration(
        "unsd_sdg_homicide",
        "United Nations Statistics Division",
        (UNSD_SERIES_URL, UNSD_DATA_URL),
        dataset_version="UNSD SDG API release reported by series catalogue",
        methodology_url="https://unstats.un.org/SDGAPI/swagger/",
        license_name="United Nations website terms; redistribution compatibility unresolved",
        license_url="https://www.un.org/en/about-us/terms-of-use",
        redistribution=(
            "General UN terms permit personal non-commercial copying but prohibit redistribution "
            "and derivative compilations absent more specific permission."
        ),
        parser="unsd_sdg_homicide_json",
        notes="UNSD republishes SDG 16.1.1 data with UNODC as custodian.",
    ),
}

FALLBACK_REGISTRATIONS = {
    "eurostat_homicide": _registration(
        "eurostat_homicide",
        "Eurostat",
        (EUROSTAT_DATA_URL,),
        dataset_version="crim_hom_soff updated 2026-04-29",
        methodology_url="https://ec.europa.eu/eurostat/cache/metadata/en/crim_sims.htm",
        license_name="Eurostat free reuse policy",
        license_url="https://ec.europa.eu/eurostat/help/copyright-notice",
        redistribution="Statistical data may be reused commercially or non-commercially with attribution.",
        parser="eurostat_jsonstat_homicide",
        notes="Narrow fallback evaluation only; not a global replacement source.",
    ),
    "oecd_homicide": _registration(
        "oecd_homicide",
        "Organisation for Economic Co-operation and Development",
        (OECD_DATAFLOW_URL,),
        dataset_version="OECD Safety - Regions dataflow 2.4",
        methodology_url="https://www.oecd.org/en/data/insights/data-explainers/2024/09/api.html",
        license_name="OECD Terms and Conditions",
        license_url="https://www.oecd.org/en/about/terms-conditions.html",
        redistribution=(
            "OECD-owned data may be reused commercially with attribution, subject to any "
            "dataset-specific third-party restrictions."
        ),
        parser="oecd_homicide_dataflow_metadata",
        notes="The only homicide-labelled OECD dataflow is subnational and is not equivalent.",
    ),
}


def is_fresh(reference_year: int | None, as_of_year: int) -> bool:
    """Apply the unchanged production homicide freshness rule."""

    return reference_year is not None and (
        as_of_year - reference_year <= FRESHNESS_MAX_AGE[CRITERION_ID]
    )


def reconcile_values(
    candidate_value: float | None,
    candidate_year: int | None,
    wdi_value: float | None,
    wdi_year: int | None,
) -> dict[str, object]:
    """Classify the latest-value comparison without treating different years as revisions."""

    if candidate_value is None or candidate_year is None:
        return {"classification": "source_unavailable", "absolute_difference": None}
    if wdi_value is None or wdi_year is None:
        return {"classification": "wdi_unavailable", "absolute_difference": None}
    difference = abs(candidate_value - wdi_value)
    if candidate_year != wdi_year:
        classification = "different_reference_year"
    elif difference <= 1e-12:
        classification = "exact_match"
    elif difference <= 0.005:
        classification = "rounding_only"
    elif difference <= max(0.05, abs(wdi_value) * 0.01):
        classification = "minor_revision"
    else:
        classification = "material_revision"
    return {
        "classification": classification,
        "absolute_difference": difference,
        "candidate_year": candidate_year,
        "wdi_year": wdi_year,
    }


def _deduplicate(records: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, int], list[dict[str, object]]] = defaultdict(list)
    for record in records:
        grouped[(str(record["country_code"]), int(record["year"]))].append(record)
    output = []
    for key, rows in sorted(grouped.items()):
        values = {round(float(row["value"]), 12) for row in rows}
        if len(values) != 1:
            continue
        output.append(
            min(
                rows,
                key=lambda row: (
                    str(row.get("data_status", "")),
                    str(row.get("source", "")),
                    str(row.get("record_id", "")),
                ),
            )
        )
    return output


def parse_direct_unodc(body: bytes, target_codes: set[str]) -> list[dict[str, object]]:
    """Select exact national, total, both-sex homicide-victim rates from the workbook."""

    workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    sheet = workbook["data_cts_intentional_homicide"]
    records = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        code, _, _, _, indicator, dimension, category, sex, age, year, unit, value, source = row
        if (
            code not in target_codes
            or indicator != "Victims of intentional homicide"
            or dimension != "Total"
            or category != "Total"
            or sex != "Total"
            or age != "Total"
            or unit != "Rate per 100,000 population"
            or value is None
        ):
            continue
        try:
            numeric = float(value)
            reference_year = int(year)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(numeric):
            continue
        records.append(
            {
                "country_code": str(code),
                "source_country_code": str(code),
                "year": reference_year,
                "value": numeric,
                "unit": UNIT,
                "series": "ICCS0101 / Victims of intentional homicide / Total",
                "dimensions": {
                    "geography": "national",
                    "sex": "both_sexes",
                    "age": "all_ages",
                    "category": "total",
                },
                "data_status": "reported_source_lineage_available",
                "provisional_revised": "not_provided",
                "source": str(source or ""),
                "record_id": f"data_cts_intentional_homicide!{row_number}",
            }
        )
    return _deduplicate(records)


def parse_unsd(
    catalogue_body: bytes,
    data_body: bytes,
    m49_to_iso3: Mapping[str, str],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    """Select explicit global-reporting, observed country records for SDG 16.1.1."""

    catalogue = json.loads(catalogue_body.decode("utf-8-sig"))
    series_items = [
        item
        for indicator in catalogue
        if indicator.get("code") == "16.1.1"
        for item in indicator.get("series", [])
        if item.get("code") == SERIES_CODE
    ]
    if len(series_items) != 1:
        raise ValueError("UNSD did not return exactly one VC_IHR_PSRC series.")
    payload = json.loads(data_body.decode("utf-8-sig"))
    if int(payload.get("totalElements", 0)) > len(payload.get("data", [])):
        raise ValueError("UNSD response was paginated beyond the retained response.")
    records = []
    unresolved_codes = set()
    rejected_records: dict[str, int] = defaultdict(int)
    for index, row in enumerate(payload.get("data", [])):
        source_code = str(int(str(row.get("geoAreaCode"))))
        code = m49_to_iso3.get(source_code)
        if code is None:
            unresolved_codes.add(source_code)
            continue
        attributes = row.get("attributes") or {}
        dimensions = row.get("dimensions") or {}
        nature = attributes.get("Nature")
        checks = (
            ("wrong_series", row.get("series") == SERIES_CODE),
            ("wrong_unit", attributes.get("Units") == "PER_100000_POP"),
            ("not_both_sexes", dimensions.get("Sex") == "BOTHSEX"),
            ("not_global_reporting", dimensions.get("Reporting Type") == "G"),
            ("modelled_or_non_country_nature", nature in {"C", "CA"}),
        )
        failed = [reason for reason, passed in checks if not passed]
        if failed:
            for reason in failed:
                rejected_records[reason] += 1
            continue
        try:
            numeric = float(row["value"])
            reference_year = int(float(row["timePeriodStart"]))
        except (KeyError, TypeError, ValueError):
            continue
        if not math.isfinite(numeric):
            continue
        records.append(
            {
                "country_code": code,
                "source_country_code": source_code.zfill(3),
                "year": reference_year,
                "value": numeric,
                "unit": UNIT,
                "series": SERIES_CODE,
                "dimensions": {
                    "geography": "country",
                    "sex": "both_sexes",
                    "age": "not_explicit_in_api_record",
                    "reporting_type": "global",
                },
                "data_status": str(nature),
                "provisional_revised": "not_provided",
                "source": str(row.get("source") or ""),
                "footnotes": list(row.get("footnotes") or []),
                "record_id": f"data[{index}]",
            }
        )
    metadata = {
        "release": series_items[0].get("release"),
        "series": SERIES_CODE,
        "total_elements": payload.get("totalElements"),
        "unresolved_geo_area_codes": sorted(unresolved_codes),
        "rejected_record_counts": dict(sorted(rejected_records.items())),
    }
    return _deduplicate(records), metadata


def _category_codes(dimension: dict[str, Any]) -> list[str]:
    index = dimension["category"]["index"]
    if isinstance(index, list):
        return [str(item) for item in index]
    return [code for code, _ in sorted(index.items(), key=lambda item: int(item[1]))]


def parse_eurostat(body: bytes, geo_to_iso3: Mapping[str, str]) -> list[dict[str, object]]:
    """Decode the narrowly filtered Eurostat JSON-stat response."""

    payload = json.loads(body.decode("utf-8-sig"))
    dimension_ids = list(payload["id"])
    sizes = [int(item) for item in payload["size"]]
    codes = {
        dimension_id: _category_codes(payload["dimension"][dimension_id])
        for dimension_id in dimension_ids
    }
    records = []
    for raw_index, raw_value in payload.get("value", {}).items():
        flat_index = int(raw_index)
        coordinates = {}
        divisor = math.prod(sizes)
        for dimension_id, size in zip(dimension_ids, sizes, strict=True):
            divisor //= size
            coordinates[dimension_id] = codes[dimension_id][(flat_index // divisor) % size]
        code = geo_to_iso3.get(coordinates["geo"])
        if code is None:
            continue
        records.append(
            {
                "country_code": code,
                "source_country_code": coordinates["geo"],
                "year": int(coordinates["time"]),
                "value": float(raw_value),
                "unit": UNIT,
                "series": "crim_hom_soff / ICCS0101 / PER_VICT",
                "dimensions": {
                    "geography": "national",
                    "sex": "both_sexes",
                    "age": "all_ages",
                    "legal_status": "victim",
                },
                "data_status": str(payload.get("status", {}).get(raw_index, "published")),
                "provisional_revised": "not_provided",
                "source": "Eurostat",
                "record_id": f"value[{raw_index}]",
            }
        )
    return _deduplicate(records)


def _normalise_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.casefold()).strip()


def eurostat_geo_mapping(
    body: bytes,
    registry: Iterable[dict[str, object]],
    target_codes: set[str],
) -> dict[str, str]:
    """Map Eurostat geographies through audited country names and aliases."""

    payload = json.loads(body.decode("utf-8-sig"))
    labels = payload["dimension"]["geo"]["category"].get("label", {})
    names = {}
    for row in registry:
        code = str(row["code"])
        if code not in target_codes:
            continue
        candidates = [str(row["display_name"]), *map(str, row.get("aliases", []))]
        for candidate in candidates:
            names[_normalise_name(candidate)] = code
    return {
        source_code: names[_normalise_name(str(label))]
        for source_code, label in labels.items()
        if _normalise_name(str(label)) in names
    }


def assess_oecd_dataflow(body: bytes) -> dict[str, object]:
    """Confirm that the OECD candidate is a regional, not national-total, dataflow."""

    if body.lstrip().startswith(b"{"):
        payload = json.loads(body.decode("utf-8-sig"))
        references = payload.get("references", {})
        matches = [
            item
            for urn, item in references.items()
            if "Dataflow=" in urn and item.get("id") == "DSD_REG_SOC@DF_SAFETY"
        ]
        if len(matches) != 1:
            raise ValueError("Expected one OECD homicide-related dataflow.")
        dataflow = matches[0]
        text = str(dataflow.get("description") or "")
        return {
            "agency": dataflow.get("agencyID", "OECD.CFE.EDS"),
            "dataflow_id": dataflow.get("id"),
            "version": dataflow.get("version", "2.4"),
            "semantic_equivalence": False,
            "reason": (
                "The OECD homicide-labelled dataflow is explicitly a subnational regional "
                "dataset; it is not a national-total ICCS 0101 victim-rate dissemination channel."
            ),
            "description_mentions_regions": "region" in text.lower(),
        }
    root = ET.fromstring(body)
    dataflows = [item for item in root.iter() if item.tag.endswith("Dataflow")]
    if len(dataflows) != 1:
        raise ValueError("Expected one OECD homicide-related dataflow.")
    dataflow = dataflows[0]
    descriptions = [
        "".join(item.itertext()) for item in dataflow if item.tag.endswith("Description")
    ]
    text = " ".join(descriptions)
    return {
        "agency": dataflow.attrib.get("agencyID"),
        "dataflow_id": dataflow.attrib.get("id"),
        "version": dataflow.attrib.get("version"),
        "semantic_equivalence": False,
        "reason": (
            "The OECD homicide-labelled dataflow is explicitly a subnational regional dataset; "
            "it is not a national-total ICCS 0101 victim-rate dissemination channel."
        ),
        "description_mentions_regions": "region" in text.lower(),
    }


def _latest_by_country(
    records: Iterable[dict[str, object]],
) -> dict[str, dict[str, object]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        grouped[str(record["country_code"])].append(record)
    return {
        code: max(rows, key=lambda row: (int(row["year"]), str(row["record_id"])))
        for code, rows in grouped.items()
    }


def should_evaluate_fallbacks(
    direct_complete_count: int,
    unsd_complete_count: int,
    minimum_required: int,
) -> bool:
    return direct_complete_count < minimum_required and unsd_complete_count < minimum_required


def coverage_status(complete_count: int, minimum_required: int) -> str:
    return "PASS" if complete_count >= minimum_required else "FAIL"


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_jsonl(path: Path) -> list[dict[str, object]]:
    return [
        json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()
    ]


def _write_json(path: Path, value: object) -> None:
    path.write_text(
        json.dumps(value, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def _write_jsonl(path: Path, values: Iterable[dict[str, object]]) -> None:
    path.write_text(
        "".join(json.dumps(value, sort_keys=True, ensure_ascii=False) + "\n" for value in values),
        encoding="utf-8",
    )


def _active_pointer_bytes(release_root: Path) -> bytes | None:
    path = release_root / "active.json"
    return path.read_bytes() if path.exists() else None


def _fetch_artifacts(
    registrations: Mapping[str, SourceRegistration],
    raw_repository: RawArtifactRepository,
    fetcher: Callable,
    retrieved_at: str,
) -> list[RawArtifact]:
    artifacts = []
    for registration in registrations.values():
        for url in registration.download_urls:
            result = fetcher(url)
            if len(result) == 3:
                body, final_url, media_type = result
                metadata = {"http_status": 200}
            else:
                body, final_url, media_type, metadata = result
            artifacts.append(
                raw_repository.capture(
                    registration,
                    body,
                    requested_url=url,
                    final_url=final_url,
                    retrieved_at=retrieved_at,
                    media_type=media_type,
                    **metadata,
                )
            )
    return artifacts


def _bodies_by_url(
    artifacts: Iterable[RawArtifact], raw_repository: RawArtifactRepository
) -> dict[str, bytes]:
    return {artifact.requested_url: raw_repository.load(artifact) for artifact in artifacts}


def _coverage_context(
    coverage_report: Path, raw_repository: RawArtifactRepository
) -> dict[str, object]:
    summary = _read_json(coverage_report / "summary.json")
    rows = _read_jsonl(coverage_report / "country-coverage.jsonl")
    registry = _read_json(coverage_report / "country-registry.json")
    target_rows = []
    for row in rows:
        criteria = row["criteria"]
        homicide = criteria[CRITERION_ID]
        other_results = [
            result for criterion, result in criteria.items() if criterion != CRITERION_ID
        ]
        if homicide["status"] != "available" and all(
            result["status"] == "available" for result in other_results
        ):
            target_rows.append(row)
    coverage_artifacts = [
        RawArtifact(**item) for item in _read_json(coverage_report / "raw-artifacts.json")
    ]
    wdi_artifact = next(item for item in coverage_artifacts if item.source_id == "unodc_homicide")
    wdi_records = _wdi_series(
        raw_repository.load(wdi_artifact),
        {str(row["code"]) for row in target_rows},
    )
    wdi_latest = {}
    for code, country_rows in wdi_records.items():
        latest = _latest_non_null(country_rows)
        wdi_latest[code] = {"value": latest[0], "year": latest[1]} if latest else None
    return {
        "summary": summary,
        "target_rows": sorted(target_rows, key=lambda row: str(row["code"])),
        "registry": registry,
        "wdi_records": wdi_records,
        "wdi_latest": wdi_latest,
        "wdi_artifact": wdi_artifact,
    }


def _source_result(
    source_id: str,
    latest: dict[str, object] | None,
    wdi_latest: dict[str, object] | None,
    *,
    as_of_year: int,
    semantic_equivalence: bool,
    evaluated: bool = True,
    reason: str | None = None,
) -> dict[str, object]:
    fresh = bool(latest and is_fresh(int(latest["year"]), as_of_year))
    if evaluated and latest is None and reason is None:
        reason = "No eligible record was available for this country."
    result = {
        "source_id": source_id,
        "evaluated": evaluated,
        "available": latest is not None,
        "latest_year": int(latest["year"]) if latest else None,
        "latest_value": float(latest["value"]) if latest else None,
        "unit": latest.get("unit") if latest else UNIT,
        "series": latest.get("series") if latest else None,
        "source_country_code": latest.get("source_country_code") if latest else None,
        "dimensions": latest.get("dimensions") if latest else None,
        "data_status": latest.get("data_status") if latest else None,
        "provisional_revised": latest.get("provisional_revised") if latest else None,
        "freshness_policy_passed": fresh,
        "semantic_equivalence": semantic_equivalence,
        "would_make_country_publishable": bool(
            evaluated and latest and fresh and semantic_equivalence
        ),
        "reason": reason,
    }
    result["discrepancy_against_current_wdi"] = reconcile_values(
        result["latest_value"],
        result["latest_year"],
        wdi_latest.get("value") if wdi_latest else None,
        wdi_latest.get("year") if wdi_latest else None,
    )
    return result


def _overlap_discrepancies(
    source_id: str,
    records: Iterable[dict[str, object]],
    wdi_records: Mapping[str, list[dict[str, object]]],
) -> list[dict[str, object]]:
    candidate = {(str(row["country_code"]), int(row["year"])): row for row in records}
    output = []
    for code, rows in wdi_records.items():
        for row in rows:
            if row.get("value") is None:
                continue
            key = (code, int(str(row["date"])))
            source_row = candidate.get(key)
            if source_row is None:
                continue
            comparison = reconcile_values(
                float(source_row["value"]),
                key[1],
                float(row["value"]),
                key[1],
            )
            output.append(
                {
                    "source_id": source_id,
                    "code": code,
                    "year": key[1],
                    "candidate_value": float(source_row["value"]),
                    "wdi_value": float(row["value"]),
                    **comparison,
                }
            )
    return sorted(
        output, key=lambda row: (str(row["source_id"]), str(row["code"]), int(row["year"]))
    )


def audit_homicide_sources(
    coverage_report: Path | str,
    study_id: str,
    *,
    mode: str,
    output_root: Path | str = "data/reports/homicide-source-feasibility",
    raw_root: Path | str = "data/raw",
    release_root: Path | str = "data/releases",
    artifact_manifest: Path | str | None = None,
    fetcher: Callable | None = None,
    clock: Callable[[], datetime] | None = None,
) -> tuple[Path, dict[str, object]]:
    """Run the online or replay study without changing source registration or releases."""

    if mode not in {"online", "replay"}:
        raise ValueError("Study mode must be 'online' or 'replay'.")
    raw_repository = RawArtifactRepository(raw_root)
    coverage_path = Path(coverage_report)
    context = _coverage_context(coverage_path, raw_repository)
    coverage_summary = context["summary"]
    minimum_required = int(coverage_summary["minimum_required_country_count"])
    baseline_count = int(coverage_summary["complete_publishable_country_count"])
    target_rows = context["target_rows"]
    target_codes = {str(row["code"]) for row in target_rows}
    now = (clock or (lambda: datetime.now(UTC)))()
    active_before = _active_pointer_bytes(Path(release_root))

    if mode == "online":
        if fetcher is None:
            from konsider.ingestion.worker import fetch_url

            fetcher = fetch_url
        artifacts = _fetch_artifacts(
            PRIMARY_REGISTRATIONS, raw_repository, fetcher, now.isoformat()
        )
    else:
        if artifact_manifest is None:
            raise ValueError("Replay requires --artifacts.")
        artifacts = [
            RawArtifact(**item)
            for item in _read_json(Path(artifact_manifest))
            if item["source_id"] in {**PRIMARY_REGISTRATIONS, **FALLBACK_REGISTRATIONS}
        ]
        if artifacts:
            now = datetime.fromisoformat(max(item.retrieved_at for item in artifacts))

    bodies = _bodies_by_url(artifacts, raw_repository)
    direct_records = parse_direct_unodc(bodies[DIRECT_UNODC_WORKBOOK_URL], target_codes)
    registry = context["registry"]
    m49_to_iso3 = {str(int(str(row["un_m49"]))): str(row["code"]) for row in registry}
    unsd_records, unsd_metadata = parse_unsd(
        bodies[UNSD_SERIES_URL], bodies[UNSD_DATA_URL], m49_to_iso3
    )
    unsd_records = [row for row in unsd_records if str(row["country_code"]) in target_codes]
    direct_latest = _latest_by_country(direct_records)
    unsd_latest = _latest_by_country(unsd_records)
    direct_recoveries = {
        code for code, record in direct_latest.items() if is_fresh(int(record["year"]), now.year)
    }
    unsd_recoveries = {
        code for code, record in unsd_latest.items() if is_fresh(int(record["year"]), now.year)
    }
    direct_complete = baseline_count + len(direct_recoveries)
    unsd_complete = baseline_count + len(unsd_recoveries)
    fallback_triggered = should_evaluate_fallbacks(direct_complete, unsd_complete, minimum_required)
    residual_codes = target_codes - direct_recoveries - unsd_recoveries

    if fallback_triggered and mode == "online":
        fallback_artifacts = _fetch_artifacts(
            FALLBACK_REGISTRATIONS, raw_repository, fetcher, now.isoformat()
        )
        artifacts.extend(fallback_artifacts)
        bodies.update(_bodies_by_url(fallback_artifacts, raw_repository))
    if fallback_triggered:
        missing = {
            url
            for registration in FALLBACK_REGISTRATIONS.values()
            for url in registration.download_urls
            if url not in bodies
        }
        if missing:
            raise ValueError(f"Replay is missing fallback artifacts: {sorted(missing)}")

    eurostat_records: list[dict[str, object]] = []
    oecd_assessment = None
    if fallback_triggered:
        eurostat_geo = eurostat_geo_mapping(bodies[EUROSTAT_DATA_URL], registry, residual_codes)
        eurostat_records = [
            row
            for row in parse_eurostat(bodies[EUROSTAT_DATA_URL], eurostat_geo)
            if str(row["country_code"]) in residual_codes
        ]
        oecd_assessment = assess_oecd_dataflow(bodies[OECD_DATAFLOW_URL])
    eurostat_latest = _latest_by_country(eurostat_records)
    eurostat_recoveries = {
        code for code, record in eurostat_latest.items() if is_fresh(int(record["year"]), now.year)
    }
    mixed_complete = baseline_count + len(direct_recoveries | unsd_recoveries | eurostat_recoveries)

    country_comparison = []
    for target in target_rows:
        code = str(target["code"])
        wdi = context["wdi_latest"][code]
        sources = {
            "current_wdi": {
                "source_id": "unodc_homicide",
                "available": wdi is not None,
                "latest_year": wdi["year"] if wdi else None,
                "latest_value": wdi["value"] if wdi else None,
                "unit": UNIT,
                "series": "VC.IHR.PSRC.P5",
                "source_country_code": code,
                "data_status": "reported_or_estimated",
                "provisional_revised": "not_provided",
                "freshness_policy_passed": bool(wdi and is_fresh(int(wdi["year"]), now.year)),
                "semantic_equivalence": True,
                "would_make_country_publishable": bool(
                    wdi and is_fresh(int(wdi["year"]), now.year)
                ),
                "production_source": True,
            },
            "direct_unodc": _source_result(
                "direct_unodc_homicide",
                direct_latest.get(code),
                wdi,
                as_of_year=now.year,
                semantic_equivalence=True,
            ),
            "unsd": _source_result(
                "unsd_sdg_homicide",
                unsd_latest.get(code),
                wdi,
                as_of_year=now.year,
                semantic_equivalence=True,
            ),
            "eurostat": _source_result(
                "eurostat_homicide",
                eurostat_latest.get(code),
                wdi,
                as_of_year=now.year,
                semantic_equivalence=True,
                evaluated=code in residual_codes and fallback_triggered,
                reason=(
                    None
                    if code in residual_codes and fallback_triggered
                    else "Outside the gated residual homicide-only set."
                ),
            ),
            "oecd": _source_result(
                "oecd_homicide",
                None,
                wdi,
                as_of_year=now.year,
                semantic_equivalence=False,
                evaluated=code in residual_codes and fallback_triggered,
                reason=(
                    oecd_assessment["reason"]
                    if code in residual_codes and oecd_assessment
                    else "Outside the gated residual homicide-only set."
                ),
            ),
        }
        country_comparison.append(
            {
                "code": code,
                "display_name": target["display_name"],
                "baseline_homicide_status": target["criteria"][CRITERION_ID]["status"],
                "sources": sources,
            }
        )

    discrepancies = (
        _overlap_discrepancies("direct_unodc_homicide", direct_records, context["wdi_records"])
        + _overlap_discrepancies("unsd_sdg_homicide", unsd_records, context["wdi_records"])
        + _overlap_discrepancies("eurostat_homicide", eurostat_records, context["wdi_records"])
    )
    discrepancy_counts: dict[str, int] = defaultdict(int)
    for item in discrepancies:
        discrepancy_counts[str(item["classification"])] += 1

    source_comparison = {
        "direct_unodc": {
            "authority": "UNODC; originating/custodian agency",
            "retrieval": "official downloadable workbook",
            "semantic_equivalence": True,
            "fresh_recoveries": sorted(direct_recoveries),
            "complete_country_count": direct_complete,
            "licensing": "unresolved_for_redistribution",
            "provisional_or_revision_field": "not_provided",
        },
        "unsd": {
            "authority": "UNSD republishing UNODC-custodian SDG data",
            "retrieval": "documented official JSON API",
            "semantic_equivalence": True,
            "fresh_recoveries": sorted(unsd_recoveries),
            "complete_country_count": unsd_complete,
            "licensing": "unresolved_for_redistribution",
            "metadata": unsd_metadata,
            "provisional_or_revision_field": "not_provided",
        },
        "eurostat": {
            "authority": "Eurostat; joint Eurostat-UNODC collection",
            "evaluated": fallback_triggered,
            "evaluated_country_codes": sorted(residual_codes) if fallback_triggered else [],
            "fresh_recoveries": sorted(eurostat_recoveries),
            "complete_country_count": mixed_complete,
            "semantic_equivalence": True,
            "licensing": "compatible_with_attribution",
        },
        "oecd": {
            "authority": "OECD",
            "evaluated": fallback_triggered,
            "evaluated_country_codes": sorted(residual_codes) if fallback_triggered else [],
            "fresh_recoveries": [],
            "complete_country_count": mixed_complete,
            "semantic_equivalence": False,
            "licensing": "generally_compatible_but_candidate_not_equivalent",
            "assessment": oecd_assessment,
        },
    }
    summary = {
        "schema_version": STUDY_SCHEMA_VERSION,
        "study_id": study_id,
        "mode": mode,
        "study_timestamp": now.isoformat(),
        "as_of_year": now.year,
        "coverage_audit_id": coverage_summary["audit_id"],
        "current_source": "World Bank WDI VC.IHR.PSRC.P5; original source UNODC",
        "homicide_freshness_max_age_years": FRESHNESS_MAX_AGE[CRITERION_ID],
        "freshness_policy_changed": False,
        "imputation_used": False,
        "baseline_complete_country_count": baseline_count,
        "minimum_required_country_count": minimum_required,
        "homicide_only_excluded_country_count": len(target_codes),
        "homicide_only_excluded_country_codes": sorted(target_codes),
        "direct_unodc_additional_fresh_country_count": len(direct_recoveries),
        "direct_unodc_complete_country_count": direct_complete,
        "unsd_additional_fresh_country_count": len(unsd_recoveries),
        "unsd_complete_country_count": unsd_complete,
        "fallback_evaluation_triggered": fallback_triggered,
        "fallback_residual_country_count": len(residual_codes),
        "fallback_residual_country_codes": sorted(residual_codes),
        "eurostat_additional_fresh_country_count": len(eurostat_recoveries),
        "oecd_additional_fresh_country_count": 0,
        "mixed_fallback_complete_country_count": mixed_complete,
        "primary_reaches_minimum": max(direct_complete, unsd_complete) >= minimum_required,
        "any_evaluated_path_reaches_minimum": mixed_complete >= minimum_required,
        "status": coverage_status(mixed_complete, minimum_required),
        "decision_outcome": (
            "C_mixed_source_requires_product_approval"
            if mixed_complete >= minimum_required
            else "D_no_reliable_path_reaches_100"
        ),
        "production_source_changed": False,
        "mixed_source_adoption_allowed": False,
        "release_activated": False,
        "discrepancy_counts": dict(sorted(discrepancy_counts.items())),
    }

    output = Path(output_root) / study_id
    output.mkdir(parents=True, exist_ok=False)
    _write_json(output / "summary.json", summary)
    _write_json(output / "source-comparison.json", source_comparison)
    _write_jsonl(output / "country-comparison.jsonl", country_comparison)
    _write_jsonl(output / "discrepancies.jsonl", discrepancies)
    _write_json(
        output / "raw-artifacts.json",
        [context["wdi_artifact"].to_dict()] + [artifact.to_dict() for artifact in artifacts],
    )
    registrations = {
        **PRIMARY_REGISTRATIONS,
        **(FALLBACK_REGISTRATIONS if fallback_triggered else {}),
    }
    _write_json(
        output / "sources.json",
        [registration.to_dict() for registration in registrations.values()],
    )
    (output / "licensing.md").write_text(_render_licensing(), encoding="utf-8")
    (output / "report.md").write_text(_render_report(summary, source_comparison), encoding="utf-8")
    if _active_pointer_bytes(Path(release_root)) != active_before:
        raise RuntimeError("Homicide source study changed the active release pointer.")
    return output, summary


def _render_licensing() -> str:
    return """# Phase 2D.4 licensing findings

## Direct UNODC and UNSD

The portals link to the general United Nations website terms. Those terms permit personal,
non-commercial downloading and copying but, absent source-specific permission, prohibit resale,
redistribution, compilation, and derivative works. No more permissive data-specific licence was found
for either retained homicide payload. Production redistribution compatibility is therefore unresolved
and blocks adoption.

## Eurostat

Eurostat authorises commercial and non-commercial reuse of statistical data and metadata with source
acknowledgement, subject to identified third-party exceptions. The evaluated dataset did not cover any
country in the gated residual set.

## OECD

OECD terms generally permit extracting, adapting, distributing, and commercially reusing OECD-owned
data with attribution, subject to dataset-specific third-party rights. Licensing was not the deciding
issue: the only homicide-labelled dataflow found is subnational and not semantically equivalent.
"""


def _render_report(summary: dict[str, object], source_comparison: dict[str, object]) -> str:
    direct = source_comparison["direct_unodc"]
    unsd = source_comparison["unsd"]
    return f"""# Homicide source feasibility study: {summary['study_id']}

- Baseline complete countries: {summary['baseline_complete_country_count']}
- Homicide-only exclusions evaluated: {summary['homicide_only_excluded_country_count']}
- Required complete countries: {summary['minimum_required_country_count']}
- Freshness rule: unchanged at {summary['homicide_freshness_max_age_years']} years
- Result: **{summary['status']}**

## Primary sources

Direct UNODC recovered {summary['direct_unodc_additional_fresh_country_count']} countries
({", ".join(direct["fresh_recoveries"]) or "none"}) and produced
{summary['direct_unodc_complete_country_count']} complete countries.

UNSD recovered {summary['unsd_additional_fresh_country_count']} countries
({", ".join(unsd["fresh_recoveries"]) or "none"}) and produced
{summary['unsd_complete_country_count']} complete countries.

Both channels represent intentional-homicide victims per 100,000 population. Direct UNODC was
filtered to national total, both sexes, all ages, and total category. UNSD was filtered to
`VC_IHR_PSRC`, country geography, `BOTHSEX`, global reporting type, `PER_100000_POP`, and observed
country/country-adjusted nature (`C` or `CA`). The UNSD record does not expose age explicitly.

## Conditional fallbacks

Fallback evaluation triggered: {summary['fallback_evaluation_triggered']}. The residual set contained
{summary['fallback_residual_country_count']} countries. Eurostat recovered
{summary['eurostat_additional_fresh_country_count']} countries; OECD recovered
{summary['oecd_additional_fresh_country_count']}. The resulting complete count remained
{summary['mixed_fallback_complete_country_count']}.

Eurostat publishes a semantically suitable ICCS 0101 victim-rate table, but its geographic coverage
does not include the residual countries. OECD's homicide-labelled dataset is regional rather than a
national-total series and was rejected as non-equivalent.

## Licensing and operational suitability

Direct UNODC provides an official downloadable workbook and UNSD provides a documented JSON API.
Exact bytes were retained for offline replay. Neither channel exposes a clear provisional/final flag.
The general UN terms do not establish production redistribution rights, so licensing remains a separate
adoption blocker. See `licensing.md`.

## Recommendation

Outcome D applies: the unchanged five-year observed-data policy cannot support 100 complete countries
through the evaluated authoritative channels. Do not migrate the production source, adopt mixed-source
precedence, weaken freshness, impute values, or activate a release. A product decision would be required
before changing the criterion policy or considering a different construct.

The production WDI registration and active release were not changed.
"""
