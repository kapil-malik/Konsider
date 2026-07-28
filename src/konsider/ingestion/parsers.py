"""Isolated parsers with record-level provenance for current and historical source families."""

from __future__ import annotations

import csv
import hashlib
import io
import json
from collections import defaultdict
from collections.abc import Iterable

from openpyxl import load_workbook

from konsider.ingestion.countries import COUNTRIES, COUNTRY_ALIASES
from konsider.ingestion.models import (
    MetricObservation,
    ObservationComponent,
    RawArtifact,
    SourceRecordReference,
)


def _observation_id(source: str, country: str, metric: str, year: int) -> str:
    return "obs_" + hashlib.sha256(f"{source}|{country}|{metric}|{year}".encode()).hexdigest()[:20]


def _annual_observation(
    *,
    records: tuple[SourceRecordReference, ...],
    source_id: str,
    country: str,
    metric: str,
    value: float,
    unit: str,
    year: int,
    observation_type: str,
    parser_version: str,
    method: str,
    flags: tuple[str, ...] = (),
    lower: float | None = None,
    upper: float | None = None,
    components: tuple[ObservationComponent, ...] = (),
) -> MetricObservation:
    artifact_ids = tuple(dict.fromkeys(record.artifact_id for record in records))
    return MetricObservation(
        observation_id=_observation_id(source_id, country, metric, year),
        country_code=country,
        metric_id=metric,
        value=float(value),
        unit=unit,
        reference_start=f"{year}-01-01",
        reference_end=f"{year}-12-31",
        source_id=source_id,
        raw_artifact_ids=artifact_ids,
        source_records=records,
        observation_type=observation_type,
        geographic_scope="national",
        parser_version=parser_version,
        method_version=method,
        quality_flags=flags,
        lower_bound=lower,
        upper_bound=upper,
        components=components,
    )


def _json(body: bytes) -> object:
    return json.loads(body.decode("utf-8-sig"))


def _latest_by_country(
    rows: Iterable[tuple[dict[str, object], SourceRecordReference]],
    code_key: str,
    year_key: str,
) -> dict[str, tuple[dict[str, object], SourceRecordReference]]:
    latest: dict[str, tuple[dict[str, object], SourceRecordReference]] = {}
    for row, reference in rows:
        code = str(row.get(code_key, ""))
        if code not in COUNTRIES or row.get(year_key) is None:
            continue
        if code not in latest or int(row[year_key]) > int(latest[code][0][year_key]):
            latest[code] = (row, reference)
    return latest


def _who_rows(artifacts: list[RawArtifact], bodies: list[bytes]):
    for artifact, body in zip(artifacts, bodies, strict=True):
        payload = _json(body)
        for index, row in enumerate(payload["value"]):  # type: ignore[index]
            record_id = "|".join(str(row.get(key, "")) for key in ("SpatialDim", "TimeDim", "Dim1"))
            yield row, SourceRecordReference(artifact.artifact_id, f"$.value[{index}]", record_id)


def parse_who_air_quality(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    all_area_codes = {"RESIDENCEAREATYPE_ALL", "RESIDENCEAREATYPE_TOTL", "RESIDENCEAREATYPE_BTSX"}
    eligible = [
        (row, ref)
        for row, ref in _who_rows(artifacts, bodies)
        if row.get("SpatialDimType") == "COUNTRY"
        and str(row.get("Dim1", "")).upper() in all_area_codes
    ]
    latest = _latest_by_country(eligible, "SpatialDim", "TimeDim")
    return [
        _annual_observation(
            records=(ref,),
            source_id="who_air_quality",
            country=code,
            metric="ambient_pm25_population_weighted",
            value=row["NumericValue"],
            unit="micrograms_per_cubic_metre",
            year=int(row["TimeDim"]),
            observation_type="modelled",
            parser_version="who_air_quality_v2",
            method="who_air_quality_observation_v2",
            flags=("modelled_estimate", "uncertainty_interval_required"),
            lower=row.get("Low"),
            upper=row.get("High"),
        )
        for code, (row, ref) in sorted(latest.items())
        if row.get("NumericValue") is not None
    ]


def parse_who_uhc(artifacts: list[RawArtifact], bodies: list[bytes]) -> list[MetricObservation]:
    latest = _latest_by_country(_who_rows(artifacts, bodies), "SpatialDim", "TimeDim")
    return [
        _annual_observation(
            records=(ref,),
            source_id="who_uhc",
            country=code,
            metric="uhc_service_coverage_index",
            value=row["NumericValue"],
            unit="index_0_100",
            year=int(row["TimeDim"]),
            observation_type="estimated",
            parser_version="who_uhc_v2",
            method="who_uhc_observation_v2",
            flags=("population_level_not_expat_access",),
        )
        for code, (row, ref) in sorted(latest.items())
        if row.get("NumericValue") is not None
    ]


def _world_bank_rows(artifact: RawArtifact, body: bytes):
    payload = _json(body)
    if not isinstance(payload, list) or len(payload) < 2 or not isinstance(payload[1], list):
        raise ValueError("Unexpected World Bank API response")
    for index, row in enumerate(payload[1]):
        record_id = (
            f"{row.get('countryiso3code')}|{row.get('indicator', {}).get('id')}|{row.get('date')}"
        )
        yield row, SourceRecordReference(artifact.artifact_id, f"$[1][{index}]", record_id)


def parse_world_bank_homicide(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    rows = []
    for row, ref in _world_bank_rows(artifacts[0], bodies[0]):
        if row.get("value") is not None:
            rows.append(
                ({**row, "iso3": row.get("countryiso3code"), "year": int(str(row["date"]))}, ref)
            )
    latest = _latest_by_country(rows, "iso3", "year")
    return [
        _annual_observation(
            records=(ref,),
            source_id="unodc_homicide",
            country=code,
            metric="intentional_homicide_rate",
            value=row["value"],
            unit="per_100000_people",
            year=int(row["year"]),
            observation_type="reported_or_estimated",
            parser_version="world_bank_homicide_v2",
            method="world_bank_homicide_observation_v2",
            flags=("secondary_distribution", "cross_country_comparability_caution"),
        )
        for code, (row, ref) in sorted(latest.items())
    ]


def parse_world_bank_icp(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    ppp = {
        row["countryiso3code"]: (row, ref)
        for row, ref in _world_bank_rows(artifacts[0], bodies[0])
        if row.get("value") is not None
    }
    exchange = {
        row["countryiso3code"]: (row, ref)
        for row, ref in _world_bank_rows(artifacts[1], bodies[1])
        if row.get("value") is not None
    }
    output = []
    for code in sorted(set(ppp) & set(exchange) & set(COUNTRIES)):
        ppp_row, ppp_ref = ppp[code]
        exchange_row, exchange_ref = exchange[code]
        value = float(ppp_row["value"]) / float(exchange_row["value"]) * 100
        output.append(
            _annual_observation(
                records=(ppp_ref, exchange_ref),
                source_id="world_bank_icp",
                country=code,
                metric="household_consumption_price_level_us_100",
                value=value,
                unit="index_us_100",
                year=2021,
                observation_type="derived",
                parser_version="world_bank_icp_v2",
                method="world_bank_icp_pli_v2",
                flags=(
                    "derived_from_official_ppp_and_exchange_rate",
                    "national_not_city_level",
                    "not_for_precise_strict_ranking",
                ),
            )
        )
    return output


def parse_wps_index(artifacts: list[RawArtifact], bodies: list[bytes]) -> list[MetricObservation]:
    workbook = load_workbook(io.BytesIO(bodies[0]), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        for header_index, row in enumerate(rows[:30]):
            labels = [str(value).strip().lower() if value is not None else "" for value in row]
            previous = (
                [
                    str(value).strip().lower() if value is not None else ""
                    for value in rows[header_index - 1]
                ]
                if header_index
                else []
            )
            country_columns = [
                i
                for i, label in enumerate(labels)
                if label in {"country", "economy", "country name", "code"}
            ]
            score_columns = [
                i
                for i, label in enumerate(labels)
                if "index score" in label or label in {"wps index", "score"}
            ]
            score_columns.extend(
                i
                for i, label in enumerate(previous)
                if "women, peace, and security index" in label and str(row[i]).strip() == "2025"
            )
            if not country_columns or not score_columns:
                continue
            result = []
            for row_index, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
                country_value = (
                    str(values[country_columns[0]]).strip()
                    if values[country_columns[0]] is not None
                    else ""
                )
                code = (
                    country_value
                    if country_value in COUNTRIES
                    else COUNTRY_ALIASES.get(country_value)
                )
                score = values[score_columns[0]]
                if code and isinstance(score, (int, float)):
                    reference = SourceRecordReference(
                        artifacts[0].artifact_id,
                        f"{sheet.title}!R{row_index}C{score_columns[0] + 1}",
                        code,
                    )
                    result.append(
                        _annual_observation(
                            records=(reference,),
                            source_id="wps_index",
                            country=code,
                            metric="women_peace_security_index",
                            value=float(score),
                            unit="index_0_1",
                            year=2025,
                            observation_type="composite",
                            parser_version="wps_index_v2",
                            method="wps_index_observation_v2",
                            flags=(
                                "mixed_reference_years",
                                "independent_composite",
                                "possible_underlying_imputation",
                            ),
                        )
                    )
            if result:
                return sorted(result, key=lambda item: item.country_code)
    raise ValueError("Could not locate country and WPS index score columns")


def _latest_wdi_rows(artifact: RawArtifact, body: bytes):
    rows = []
    for row, reference in _world_bank_rows(artifact, body):
        if row.get("value") is None:
            continue
        normalized = {**row, "iso3": row.get("countryiso3code"), "year": int(str(row["date"]))}
        rows.append((normalized, reference))
    return _latest_by_country(rows, "iso3", "year")


def parse_world_bank_pm25(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    latest = _latest_wdi_rows(artifacts[0], bodies[0])
    return [
        _annual_observation(
            records=(ref,),
            source_id="world_bank_pm25",
            country=code,
            metric="ambient_pm25_population_weighted",
            value=row["value"],
            unit="micrograms_per_cubic_metre",
            year=int(row["year"]),
            observation_type="modelled",
            parser_version="world_bank_pm25_v1",
            method="world_bank_pm25_observation_v1",
            flags=("wdi_distribution", "modelled_estimate", "cross_country_comparison_only"),
        )
        for code, (row, ref) in sorted(latest.items())
    ]


def parse_world_bank_uhc(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    latest = _latest_wdi_rows(artifacts[0], bodies[0])
    return [
        _annual_observation(
            records=(ref,),
            source_id="world_bank_uhc",
            country=code,
            metric="uhc_service_coverage_index",
            value=row["value"],
            unit="index_0_100",
            year=int(row["year"]),
            observation_type="estimated",
            parser_version="world_bank_uhc_v1",
            method="world_bank_uhc_observation_v1",
            flags=("wdi_distribution", "population_level_not_expat_access", "upstream_latest_2021"),
        )
        for code, (row, ref) in sorted(latest.items())
    ]


def parse_world_bank_homicide_current(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    latest = _latest_wdi_rows(artifacts[0], bodies[0])
    return [
        _annual_observation(
            records=(ref,),
            source_id="unodc_homicide",
            country=code,
            metric="intentional_homicide_rate",
            value=row["value"],
            unit="per_100000_people",
            year=int(row["year"]),
            observation_type="reported_or_estimated",
            parser_version="world_bank_homicide_v3",
            method="world_bank_homicide_observation_v3",
            flags=(
                "wdi_distribution",
                "secondary_distribution",
                "cross_country_comparability_caution",
            ),
        )
        for code, (row, ref) in sorted(latest.items())
    ]


def parse_world_bank_icp_current(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    ppp_rows = list(_world_bank_rows(artifacts[0], bodies[0]))
    exchange_rows = list(_world_bank_rows(artifacts[1], bodies[1]))
    ppp = {
        (row["countryiso3code"], int(str(row["date"]))): (row, ref)
        for row, ref in ppp_rows
        if row.get("value") is not None
    }
    exchange = {
        (row["countryiso3code"], int(str(row["date"]))): (row, ref)
        for row, ref in exchange_rows
        if row.get("value") is not None
    }
    output = []
    for code in sorted(COUNTRIES):
        years = sorted(
            {year for country, year in ppp if country == code}
            & {year for country, year in exchange if country == code}
        )
        if not years:
            continue
        year = years[-1]
        ppp_row, ppp_ref = ppp[(code, year)]
        exchange_row, exchange_ref = exchange[(code, year)]
        ppp_value, exchange_value = float(ppp_row["value"]), float(exchange_row["value"])
        components = (
            ObservationComponent(
                "household_consumption_ppp",
                ppp_value,
                "lcu_per_international_dollar",
                year,
                ppp_ref,
            ),
            ObservationComponent(
                "official_exchange_rate", exchange_value, "lcu_per_us_dollar", year, exchange_ref
            ),
        )
        output.append(
            _annual_observation(
                records=(ppp_ref, exchange_ref),
                source_id="world_bank_icp",
                country=code,
                metric="household_consumption_price_level_us_100",
                value=ppp_value / exchange_value * 100,
                unit="index_us_100",
                year=year,
                observation_type="derived",
                parser_version="world_bank_icp_v3",
                method="world_bank_icp_pli_v3",
                components=components,
                flags=(
                    "wdi_distribution",
                    "derived_from_official_ppp_and_exchange_rate",
                    "national_not_city_level",
                    "broad_band_only",
                    "not_for_precise_strict_ranking",
                ),
            )
        )
    return output


def parse_world_bank_wbl(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    workbook = load_workbook(io.BytesIO(bodies[0]), read_only=True, data_only=True)
    sheet = workbook["WBL Economy Scores"]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(i for i, row in enumerate(rows[:20]) if row and row[0] == "Economy")
    headers = list(rows[header_index])
    iso_column = headers.index("ISO Code")
    year_column = headers.index("Report Year")
    score_column = headers.index("I. Economy LF Index")
    latest_rows: dict[str, tuple[int, int, tuple[object, ...]]] = {}
    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        code = str(row[iso_column]).strip() if row[iso_column] is not None else ""
        if code not in COUNTRIES or not isinstance(row[score_column], (int, float)):
            continue
        report_year = int(row[year_column])
        if code not in latest_rows or report_year > latest_rows[code][0]:
            latest_rows[code] = (report_year, row_index, row)

    output = []
    for code, (report_year, row_index, row) in sorted(latest_rows.items()):
        reference = SourceRecordReference(
            artifacts[0].artifact_id,
            f"WBL Economy Scores!R{row_index}C{score_column + 1}",
            f"{code}|WBL_LF_INDEX|report-{report_year}",
        )
        output.append(
            _annual_observation(
                records=(reference,),
                source_id="world_bank_wbl",
                country=code,
                metric="women_legal_economic_equality",
                value=float(row[score_column]),
                unit="index_0_100",
                year=2025,
                observation_type="composite",
                parser_version="world_bank_wbl_v1",
                method="wbl_2026_legal_framework_index_v1",
                flags=(
                    "world_bank_primary_dataset",
                    "de_jure_legal_framework",
                    "not_de_facto_outcomes",
                    "data_current_2025_10_01",
                ),
            )
        )
    return output


def _piecewise_0_100(value: float, anchors: tuple[tuple[float, float], ...]) -> float:
    if value <= anchors[0][0]:
        return anchors[0][1]
    if value >= anchors[-1][0]:
        return anchors[-1][1]
    for (x0, y0), (x1, y1) in zip(anchors, anchors[1:], strict=False):
        if x0 <= value <= x1:
            return y0 + (value - x0) / (x1 - x0) * (y1 - y0)
    raise AssertionError("component interpolation failed")


def parse_world_bank_infrastructure(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    latest_by_component = [
        _latest_wdi_rows(artifact, body) for artifact, body in zip(artifacts, bodies, strict=True)
    ]
    component_ids = (
        "internet_users_percent",
        "fixed_broadband_per_100",
        "lpi_infrastructure_quality",
    )
    component_units = ("percent_population", "subscriptions_per_100", "index_1_5")
    output = []
    for code in sorted(COUNTRIES):
        if not all(code in latest for latest in latest_by_component):
            continue
        rows_refs = [latest[code] for latest in latest_by_component]
        values = [float(row["value"]) for row, _ in rows_refs]
        normalized = (
            min(max(values[0], 0), 100),
            _piecewise_0_100(values[1], ((0, 0), (10, 25), (20, 50), (30, 75), (40, 100))),
            _piecewise_0_100(values[2], ((1, 0), (2, 25), (3, 50), (4, 75), (5, 100))),
        )
        references = tuple(ref for _, ref in rows_refs)
        components = tuple(
            ObservationComponent(component_id, value, unit, int(row["year"]), ref)
            for component_id, unit, value, (row, ref) in zip(
                component_ids, component_units, values, rows_refs, strict=True
            )
        )
        years = [component.reference_year for component in components]
        output.append(
            _annual_observation(
                records=references,
                source_id="world_bank_infrastructure",
                country=code,
                metric="infrastructure_readiness_composite",
                value=sum(normalized) / len(normalized),
                unit="index_0_100",
                year=max(years),
                observation_type="derived_composite",
                parser_version="world_bank_infrastructure_v1",
                method="wdi_infrastructure_equal_weight_v1",
                components=components,
                flags=(
                    "wdi_distribution",
                    "equal_weight_three_components",
                    "mixed_reference_years",
                    "digital_and_trade_transport_scope",
                    "itu_attribution_required",
                ),
            )
        )
    return output


def _parse_world_bank_wgi(
    artifacts: list[RawArtifact],
    bodies: list[bytes],
    *,
    source_id: str,
    metric_id: str,
    parser_version: str,
    sheet_name: str,
) -> list[MetricObservation]:
    workbook = load_workbook(io.BytesIO(bodies[0]), read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    code_col = headers.index("Economy (code)")
    year_col = headers.index("Year")
    estimate_col = headers.index("Governance estimate (approx. -2.5 to +2.5)")
    stderr_col = headers.index("Standard error (estimate)")
    lower_col = headers.index("Lower threshold (90% conf. int. estimate)")
    upper_col = headers.index("Upper threshold (90% conf. int. estimate)")
    latest = {}
    for row_number, row in enumerate(rows, start=2):
        code = str(row[code_col] or "").strip()
        if code not in COUNTRIES or not isinstance(row[estimate_col], (int, float)):
            continue
        year = int(row[year_col])
        if code not in latest or year > latest[code][0]:
            latest[code] = (year, row_number, row)
    output = []
    for code in sorted(COUNTRIES):
        if code not in latest:
            continue
        year, row_number, row = latest[code]
        value = float(row[estimate_col])
        reference = SourceRecordReference(
            artifacts[0].artifact_id,
            f"{sheet_name}!R{row_number}",
            f"{code}|{sheet_name}|{year}",
        )
        output.append(
            _annual_observation(
                records=(reference,),
                source_id=source_id,
                country=code,
                metric=metric_id,
                value=value,
                unit="estimate_minus_2_5_to_2_5",
                year=year,
                observation_type="perception_based_composite",
                parser_version=parser_version,
                method="wgi_estimate_with_uncertainty_v1",
                flags=(
                    "world_bank_primary_dataset",
                    "perception_based",
                    "published_standard_error_retained",
                    "broad_band_only",
                ),
                lower=float(row[lower_col]),
                upper=float(row[upper_col]),
                components=(
                    ObservationComponent(
                        "published_standard_error",
                        float(row[stderr_col]),
                        "standard_error",
                        year,
                        reference,
                    ),
                ),
            )
        )
    return output


def parse_world_bank_wgi_political_stability(artifacts, bodies):
    return _parse_world_bank_wgi(
        artifacts,
        bodies,
        source_id="world_bank_wgi_political_stability",
        metric_id="political_stability",
        parser_version="world_bank_wgi_xlsx_v1",
        sheet_name="pv",
    )


def parse_world_bank_wgi_rule_of_law(artifacts, bodies):
    return _parse_world_bank_wgi(
        artifacts,
        bodies,
        source_id="world_bank_wgi_rule_of_law",
        metric_id="rule_of_law",
        parser_version="world_bank_wgi_xlsx_v1",
        sheet_name="rl",
    )


def parse_world_bank_migrant_stock(artifacts, bodies):
    latest = _latest_wdi_rows(artifacts[0], bodies[0])
    return [
        _annual_observation(
            records=(ref,),
            source_id="world_bank_migrant_stock",
            country=code,
            metric="established_immigrant_presence",
            value=row["value"],
            unit="percent_population",
            year=int(row["year"]),
            observation_type="estimated_stock",
            parser_version="world_bank_migrant_stock_v1",
            method="wdi_migrant_stock_share_v1",
            flags=(
                "wdi_distribution",
                "un_population_division_upstream",
                "preference_property_not_universal_quality",
                "not_integration_or_visa_access",
            ),
        )
        for code, (row, ref) in sorted(latest.items())
    ]


_JOB_MARKET_REFERENCE_YEAR = 2025
_JOB_MARKET_COMPONENTS = {
    "EMP_2WAP_SEX_AGE_RT": "employment_to_population_ratio",
    "EAP_2WAP_SEX_AGE_RT": "labour_force_participation_rate",
    "UNE_2EAP_SEX_AGE_RT": "unemployment_rate",
}
_JOB_MARKET_COMPONENT_ORDER = (
    "employment_to_population_ratio",
    "labour_force_participation_rate",
    "unemployment_rate",
)


def _average_rank_scores(values: dict[str, float], *, higher_is_better: bool) -> dict[str, float]:
    ordered = sorted(values.items(), key=lambda item: (item[1], item[0]))
    if len(ordered) == 1:
        return {ordered[0][0]: 5.5}
    scores: dict[str, float] = {}
    start = 0
    while start < len(ordered):
        end = start + 1
        while end < len(ordered) and ordered[end][1] == ordered[start][1]:
            end += 1
        percentile = ((start + end - 1) / 2) / (len(ordered) - 1)
        if not higher_is_better:
            percentile = 1 - percentile
        for index in range(start, end):
            scores[ordered[index][0]] = 1 + 9 * percentile
        start = end
    return scores


def _ilostat_job_market_rows(artifacts: list[RawArtifact], bodies: list[bytes]):
    target: dict[str, dict[str, tuple[float, SourceRecordReference]]] = defaultdict(dict)
    latest_year: dict[str, dict[str, int]] = defaultdict(dict)
    for artifact, body in zip(artifacts, bodies, strict=True):
        rows = csv.DictReader(io.StringIO(body.decode("utf-8-sig")))
        for line_number, row in enumerate(rows, start=2):
            if row.get("sex") != "SEX_T" or row.get("classif1") != "AGE_YTHADULT_YGE15":
                continue
            component_id = _JOB_MARKET_COMPONENTS.get(str(row.get("indicator") or ""))
            code = str(row.get("ref_area") or "").strip()
            if component_id is None or code not in COUNTRIES:
                continue
            try:
                year = int(row["time"])
                value = float(row["obs_value"])
            except (KeyError, TypeError, ValueError):
                continue
            latest_year[code][component_id] = max(year, latest_year[code].get(component_id, year))
            if year != _JOB_MARKET_REFERENCE_YEAR or component_id in target[code]:
                continue
            reference = SourceRecordReference(
                artifact.artifact_id,
                f"{artifact.requested_url}#row={line_number}",
                f"{code}|{row['indicator']}|{year}",
            )
            target[code][component_id] = (value, reference)
    return target, latest_year


def classify_ilostat_job_market_outcomes(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> dict[str, tuple[str, tuple[str, ...]]]:
    """Classify every stable country from the pinned source bytes."""

    target, latest_year = _ilostat_job_market_rows(artifacts, bodies)
    output = {}
    for code in sorted(COUNTRIES):
        components = target.get(code, {})
        if all(component in components for component in _JOB_MARKET_COMPONENT_ORDER):
            output[code] = ("valid", ())
            continue
        latest = latest_year.get(code, {})
        if (
            all(component in latest for component in _JOB_MARKET_COMPONENT_ORDER)
            and max(latest.values()) < _JOB_MARKET_REFERENCE_YEAR
        ):
            output[code] = ("stale", ("FRS_STALE",))
            continue
        if not latest:
            output[code] = ("missing", ("COV_SOURCE_RECORD_MISSING",))
            continue
        missing = tuple(
            f"VAL_COMPONENT_MISSING:{component}"
            for component in _JOB_MARKET_COMPONENT_ORDER
            if component not in components
        )
        output[code] = ("invalid", missing or ("VAL_COMPONENT_SET_INCOMPLETE",))
    return output


def parse_ilostat_job_market_opportunity(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    """Build the frozen 2025 equal-component labour-market percentile index."""

    target, _ = _ilostat_job_market_rows(artifacts, bodies)
    valid = {
        code: components
        for code, components in target.items()
        if all(component in components for component in _JOB_MARKET_COMPONENT_ORDER)
    }
    employment = _average_rank_scores(
        {
            code: components["employment_to_population_ratio"][0]
            for code, components in valid.items()
        },
        higher_is_better=True,
    )
    participation = _average_rank_scores(
        {
            code: components["labour_force_participation_rate"][0]
            for code, components in valid.items()
        },
        higher_is_better=True,
    )
    reverse_unemployment = _average_rank_scores(
        {code: components["unemployment_rate"][0] for code, components in valid.items()},
        higher_is_better=False,
    )
    component_scores = {
        "employment_to_population_ratio": employment,
        "labour_force_participation_rate": participation,
        "unemployment_rate": reverse_unemployment,
    }
    units = {
        "employment_to_population_ratio": "percent_working_age_population",
        "labour_force_participation_rate": "percent_working_age_population",
        "unemployment_rate": "percent_labour_force",
    }
    observations = []
    for code in sorted(valid):
        components = valid[code]
        records = tuple(components[item][1] for item in _JOB_MARKET_COMPONENT_ORDER)
        value = sum(component_scores[item][code] for item in _JOB_MARKET_COMPONENT_ORDER) / 3
        observations.append(
            _annual_observation(
                records=records,
                source_id="ilostat_job_market_opportunity",
                country=code,
                metric="overall_job_market_opportunity",
                value=value,
                unit="equal_component_percentile_index_1_10",
                year=_JOB_MARKET_REFERENCE_YEAR,
                observation_type="derived_composite_of_modelled_national_estimates",
                parser_version="ilostat_job_market_opportunity_v1",
                method="ilostat_job_market_equal_component_percentiles_v1",
                components=tuple(
                    ObservationComponent(
                        component_id,
                        components[component_id][0],
                        units[component_id],
                        _JOB_MARKET_REFERENCE_YEAR,
                        components[component_id][1],
                    )
                    for component_id in _JOB_MARKET_COMPONENT_ORDER
                ),
                flags=(
                    "modelled_national_estimates",
                    "total_population_age_15_plus",
                    "equal_weight_three_components",
                    "average_rank_percentiles",
                    "unemployment_direction_reversed",
                    "occupation_neutral",
                    "future_projections_excluded",
                ),
            )
        )
    return observations


_HCI_PLUS_FRESHNESS_YEAR = 2024


def _hci_plus_schooling_rows(artifact: RawArtifact, body: bytes):
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover - production dependency guard
        raise RuntimeError("The HCI+ production parser requires pandas.") from exc

    frame = pd.read_stata(io.BytesIO(body))
    required = {"iso3c", "year", "hlo_mf", "lays_mf"}
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"HCI+ panel is missing fields: {', '.join(sorted(missing))}")
    latest: dict[str, tuple[int, float | None, float | None, SourceRecordReference]] = {}
    for index, row in frame.iterrows():
        code = str(row["iso3c"]).strip()
        if code not in COUNTRIES:
            continue
        try:
            year = int(row["year"])
        except (TypeError, ValueError):
            continue
        hlo = None if pd.isna(row["hlo_mf"]) else float(row["hlo_mf"])
        lays = None if pd.isna(row["lays_mf"]) else float(row["lays_mf"])
        if hlo is None and lays is None:
            continue
        reference = SourceRecordReference(
            artifact.artifact_id,
            f"row={int(index) + 2}",
            f"{code}|HCI_PLUS_LAYS|{year}",
        )
        if code not in latest or year > latest[code][0]:
            latest[code] = (year, hlo, lays, reference)
    return latest


def classify_hci_plus_schooling_outcomes(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> dict[str, tuple[str, tuple[str, ...]]]:
    """Classify the stable universe for the frozen LAYS construct."""

    if len(artifacts) != 1:
        raise ValueError("HCI+ schooling requires exactly one source artifact.")
    latest = _hci_plus_schooling_rows(artifacts[0], bodies[0])
    output = {}
    for code in sorted(COUNTRIES):
        row = latest.get(code)
        if row is None:
            output[code] = ("missing", ("COV_SOURCE_RECORD_MISSING",))
            continue
        year, hlo, lays, _ = row
        missing = tuple(
            reason
            for value, reason in (
                (hlo, "VAL_COMPONENT_MISSING:harmonized_learning_outcome"),
                (lays, "VAL_COMPONENT_MISSING:learning_adjusted_years_schooling"),
            )
            if value is None
        )
        if missing:
            output[code] = ("invalid", missing)
        elif year < _HCI_PLUS_FRESHNESS_YEAR:
            output[code] = ("stale", ("FRS_STALE",))
        else:
            output[code] = ("valid", ())
    return output


def parse_world_bank_hci_plus_schooling(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    """Parse fresh learning-adjusted years without requiring unrelated HCI+ components."""

    if len(artifacts) != 1:
        raise ValueError("HCI+ schooling requires exactly one source artifact.")
    latest = _hci_plus_schooling_rows(artifacts[0], bodies[0])
    observations = []
    for code, (year, hlo, lays, reference) in sorted(latest.items()):
        if year < _HCI_PLUS_FRESHNESS_YEAR or hlo is None or lays is None:
            continue
        observations.append(
            _annual_observation(
                records=(reference,),
                source_id="world_bank_hci_plus_schooling",
                country=code,
                metric="school_education_quality",
                value=lays,
                unit="learning_adjusted_years",
                year=year,
                observation_type="national_modelled_and_harmonized_outcome",
                parser_version="world_bank_hci_plus_schooling_v1",
                method="hci_plus_learning_adjusted_years_v1",
                components=(
                    ObservationComponent(
                        "harmonized_learning_outcome",
                        hlo,
                        "harmonized_test_score",
                        year,
                        reference,
                    ),
                ),
                flags=(
                    "national_modelled_outcome",
                    "harmonized_assessment_inputs",
                    "mixed_underlying_reference_periods",
                    "not_city_or_school_specific",
                    "catalogue_file_year_label_discrepancy",
                ),
            )
        )
    return observations


def parse_wipo_innovation_outputs(
    artifacts: list[RawArtifact], bodies: list[bytes]
) -> list[MetricObservation]:
    """Parse WIPO's published output sub-index without republishing input indicators."""

    if len(artifacts) != 1:
        raise ValueError("WIPO innovation outputs requires exactly one source artifact.")
    workbook = load_workbook(io.BytesIO(bodies[0]), read_only=True, data_only=True)
    sheet = workbook["Data"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    required = {"ISO3", "ECONOMY_NAME", "NAME", "SCORE", "RANK"}
    if not required.issubset(headers):
        raise ValueError(f"WIPO workbook is missing fields: {sorted(required - set(headers))}")
    columns = {name: headers.index(name) for name in required}
    output = []
    for row_number, row in enumerate(rows, start=2):
        code = str(row[columns["ISO3"]] or "").strip()
        if (
            code not in COUNTRIES
            or row[columns["NAME"]] != "Innovation outputs"
            or not isinstance(row[columns["SCORE"]], (int, float))
        ):
            continue
        reference = SourceRecordReference(
            artifacts[0].artifact_id,
            f"Data!R{row_number}C{columns['SCORE'] + 1}",
            f"{code}|GII2025|INNOVATION_OUTPUTS",
        )
        output.append(
            _annual_observation(
                records=(reference,),
                source_id="wipo_innovation_outputs",
                country=code,
                metric="research_innovation_ecosystem",
                value=float(row[columns["SCORE"]]),
                unit="wipo_innovation_outputs_score_0_100",
                year=2025,
                observation_type="published_composite_sub_index",
                parser_version="wipo_innovation_outputs_v1",
                method="wipo_gii_innovation_outputs_v1",
                flags=(
                    "published_wipo_sub_index",
                    "mixed_source_years",
                    "third_party_input_columns_not_redistributed",
                    "national_not_city_cluster",
                    "experimental",
                ),
            )
        )
    return sorted(output, key=lambda item: item.country_code)


PARSERS = {
    "who_air_quality": parse_who_air_quality,
    "who_uhc": parse_who_uhc,
    "world_bank_homicide": parse_world_bank_homicide,
    "world_bank_icp": parse_world_bank_icp,
    "wps_index": parse_wps_index,
    "world_bank_pm25": parse_world_bank_pm25,
    "world_bank_uhc": parse_world_bank_uhc,
    "world_bank_homicide_current": parse_world_bank_homicide_current,
    "world_bank_icp_current": parse_world_bank_icp_current,
    "world_bank_wbl": parse_world_bank_wbl,
    "world_bank_infrastructure": parse_world_bank_infrastructure,
    "world_bank_wgi_political_stability": parse_world_bank_wgi_political_stability,
    "world_bank_wgi_rule_of_law": parse_world_bank_wgi_rule_of_law,
    "world_bank_migrant_stock": parse_world_bank_migrant_stock,
    "ilostat_job_market_opportunity": parse_ilostat_job_market_opportunity,
    "world_bank_hci_plus_schooling": parse_world_bank_hci_plus_schooling,
    "wipo_innovation_outputs": parse_wipo_innovation_outputs,
}
