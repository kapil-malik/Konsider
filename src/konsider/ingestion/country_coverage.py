"""Reproducible country-universe discovery and complete-case coverage auditing."""

from __future__ import annotations

import io
import csv
import json
import math
import re
import zipfile
from collections import Counter
from dataclasses import dataclass, replace
from datetime import UTC, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook

from konsider.ingestion.models import RawArtifact, SourceRegistration
from konsider.ingestion.registry import SOURCES
from konsider.ingestion.scoring import CURRENT_THRESHOLD_METHODS, _algorithm_scores
from konsider.ingestion.validation import FRESHNESS_MAX_AGE, RANGES
from konsider.repositories.raw_artifact_repository import RawArtifactRepository

UN_M49_URL = "https://unstats.un.org/unsd/methodology/m49/overview/"
UN_MIGRANT_STOCK_URL = (
    "https://www.un.org/development/desa/pd/sites/www.un.org.development.desa.pd/files/"
    "undesa_pd_2024_ims_stock_by_sex_and_destination.xlsx"
)
WORLD_BANK_COUNTRY_METADATA_URL = "https://api.worldbank.org/v2/country?format=json&per_page=400"
AUDIT_SCHEMA_VERSION = "country-coverage-audit-1.0"

SUPPORTED_CRITERION_SOURCE_IDS = {
    "ambient_pm25_population_weighted": "world_bank_pm25",
    "intentional_homicide_rate": "unodc_homicide",
    "household_consumption_price_level_us_100": "world_bank_icp",
    "women_legal_economic_equality": "world_bank_wbl",
    "infrastructure_readiness_composite": "world_bank_infrastructure",
}


def enabled_criteria_from_catalog(
    catalog_path: Path | str = "data/catalogs/consumer-catalog-1.0.json",
) -> list[str]:
    catalog = json.loads(Path(catalog_path).read_text(encoding="utf-8"))
    criteria = [
        str(item["id"])
        for item in catalog["criteria"]
        if item.get("ready") and item.get("default_enabled")
    ]
    unsupported = set(criteria) - set(SUPPORTED_CRITERION_SOURCE_IDS)
    if unsupported:
        raise ValueError(
            f"Coverage audit has no implementation for enabled criteria: {sorted(unsupported)}"
        )
    registered_by_criterion = {source.criterion_id: source.source_id for source in SOURCES.values()}
    missing_sources = set(criteria) - set(registered_by_criterion)
    if missing_sources:
        raise ValueError(f"Enabled criteria have no registered source: {sorted(missing_sources)}")
    return criteria


def _universe_registration(
    source_id: str,
    url: str,
    *,
    publisher: str,
    dataset_version: str,
    parser: str,
    parser_version: str,
    license_name: str,
    license_url: str,
    attribution: str,
) -> SourceRegistration:
    return SourceRegistration(
        source_id=source_id,
        criterion_id="country_universe_selection",
        publisher=publisher,
        distributor=None,
        canonical_page_url=url,
        download_urls=(url,),
        access_method="official public download",
        pagination="none",
        dataset_version=dataset_version,
        source_version=dataset_version,
        reference_period="current published edition",
        update_frequency="periodic",
        methodology_url=url,
        license_name=license_name,
        license_url=license_url,
        redistribution="Raw bytes are retained locally; only compact derived metadata is committed.",
        permitted_usage="Used for reproducible country identity and candidate selection.",
        attribution=attribution,
        license_evidence=license_url,
        parser=parser,
        parser_version=parser_version,
        official_or_independent="official",
        notes="Discovery-only source; it never contributes to affinity scoring.",
    )


UNIVERSE_SOURCES = {
    "un_m49_country_registry": _universe_registration(
        "un_m49_country_registry",
        UN_M49_URL,
        publisher="United Nations Statistics Division",
        dataset_version="UN M49 online standard retrieved 2026-07-23",
        parser="un_m49_html_v1",
        parser_version="un_m49_html_v1",
        license_name="United Nations website terms",
        license_url="https://www.un.org/en/about-us/terms-of-use",
        attribution="United Nations Statistics Division, Standard country or area codes (M49).",
    ),
    "un_desa_migrant_stock": _universe_registration(
        "un_desa_migrant_stock",
        UN_MIGRANT_STOCK_URL,
        publisher="United Nations DESA Population Division",
        dataset_version="International Migrant Stock 2024",
        parser="un_desa_migrant_destination_v1",
        parser_version="un_desa_migrant_destination_v1",
        license_name="Creative Commons Attribution 3.0 IGO",
        license_url="https://creativecommons.org/licenses/by/3.0/igo/",
        attribution="United Nations DESA Population Division, International Migrant Stock 2024.",
    ),
    "world_bank_country_metadata": _universe_registration(
        "world_bank_country_metadata",
        WORLD_BANK_COUNTRY_METADATA_URL,
        publisher="World Bank",
        dataset_version="World Bank Country API retrieved 2026-07-23",
        parser="world_bank_country_metadata_v1",
        parser_version="world_bank_country_metadata_v1",
        license_name="Creative Commons Attribution 4.0 International",
        license_url="https://datacatalog.worldbank.org/public-licenses#cc-by",
        attribution="World Bank Country API metadata.",
    ),
}


def _candidate_country_url(url: str, country_codes: Iterable[str]) -> str:
    encoded_codes = ";".join(sorted(country_codes))
    return re.sub(r"/country/[^/]+/indicator/", f"/country/{encoded_codes}/indicator/", url)


def _candidate_urls(
    urls: Iterable[str], country_codes: Iterable[str], *, chunk_size: int = 20
) -> tuple[str, ...]:
    codes = sorted(country_codes)
    chunks = [codes[index : index + chunk_size] for index in range(0, len(codes), chunk_size)]
    output = []
    for url in urls:
        if "/country/" not in url:
            output.append(url)
            continue
        output.extend(_candidate_country_url(url, chunk) for chunk in chunks)
    return tuple(output)


def _world_bank_bulk_url(url: str) -> str:
    match = re.search(r"/indicator/([^?]+)", url)
    if not match:
        return url
    return (
        f"https://api.worldbank.org/v2/country/all/indicator/{match.group(1)}" "?downloadformat=csv"
    )


def audit_source_registrations(
    country_codes: Iterable[str], enabled_criteria: Iterable[str] | None = None
) -> dict[str, SourceRegistration]:
    """Return discovery sources plus candidate-set variants of ranking sources."""

    registrations = dict(UNIVERSE_SOURCES)
    criteria = list(enabled_criteria or enabled_criteria_from_catalog())
    registered_by_criterion = {source.criterion_id: source.source_id for source in SOURCES.values()}
    for source_id in (registered_by_criterion[criterion] for criterion in criteria):
        source = SOURCES[source_id]
        download_urls = (
            tuple(_world_bank_bulk_url(url) for url in source.download_urls)
            if all("/indicator/" in url for url in source.download_urls)
            else _candidate_urls(source.download_urls, country_codes)
        )
        registrations[source_id] = replace(
            source,
            download_urls=download_urls,
            access_method=(
                f"{source.access_method}; official indicator-level CSV ZIP for coverage audit"
                if all("downloadformat=csv" in url for url in download_urls)
                else f"{source.access_method}; complete candidate set for coverage audit"
            ),
        )
    return registrations


@dataclass(frozen=True)
class CountryRegistryRecord:
    code: str
    display_name: str
    un_m49: str
    region: str
    subregion: str
    entity_type: str
    active: bool
    source_codes: dict[str, str]
    aliases: tuple[str, ...] = ()

    def to_dict(self) -> dict[str, object]:
        return {
            "code": self.code,
            "display_name": self.display_name,
            "un_m49": self.un_m49,
            "region": self.region,
            "subregion": self.subregion,
            "entity_type": self.entity_type,
            "active": self.active,
            "source_codes": dict(sorted(self.source_codes.items())),
            "aliases": list(self.aliases),
        }


class _M49TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_table = False
        self.in_cell = False
        self.depth = 0
        self.cell_parts: list[str] = []
        self.row: list[str] = []
        self.rows: list[list[str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "table" and attributes.get("id") == "downloadTableEN":
            self.in_table = True
            self.depth = 1
            return
        if not self.in_table:
            return
        if tag == "table":
            self.depth += 1
        elif tag == "tr":
            self.row = []
        elif tag in {"td", "th"}:
            self.in_cell = True
            self.cell_parts = []

    def handle_data(self, data: str) -> None:
        if self.in_cell:
            self.cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if not self.in_table:
            return
        if tag in {"td", "th"} and self.in_cell:
            self.row.append(" ".join("".join(self.cell_parts).split()))
            self.in_cell = False
        elif tag == "tr" and self.row:
            self.rows.append(self.row)
        elif tag == "table":
            self.depth -= 1
            if self.depth == 0:
                self.in_table = False


def parse_m49_registry(body: bytes) -> dict[str, dict[str, str]]:
    parser = _M49TableParser()
    parser.feed(body.decode("utf-8-sig"))
    if not parser.rows:
        raise ValueError("Could not locate the English UN M49 download table.")
    header = parser.rows[0]
    required = {
        "Region Name",
        "Sub-region Name",
        "Country or Area",
        "M49 Code",
        "ISO-alpha2 Code",
        "ISO-alpha3 Code",
    }
    if not required <= set(header):
        raise ValueError("UN M49 table headers changed.")
    positions = {name: header.index(name) for name in required}
    records: dict[str, dict[str, str]] = {}
    for row in parser.rows[1:]:
        if len(row) < len(header):
            row += [""] * (len(header) - len(row))
        code = row[positions["ISO-alpha3 Code"]].strip().upper()
        if not re.fullmatch(r"[A-Z]{3}", code):
            continue
        if code in records:
            raise ValueError(f"Duplicate active UN M49 ISO3 code: {code}")
        records[code] = {
            "code": code,
            "display_name": row[positions["Country or Area"]].strip(),
            "un_m49": row[positions["M49 Code"]].strip().zfill(3),
            "iso2": row[positions["ISO-alpha2 Code"]].strip().upper(),
            "region": row[positions["Region Name"]].strip(),
            "subregion": row[positions["Sub-region Name"]].strip(),
        }
    if len(records) < 200:
        raise ValueError(f"UN M49 registry unexpectedly contained only {len(records)} records.")
    return records


def parse_world_bank_country_metadata(body: bytes) -> dict[str, dict[str, str]]:
    payload = json.loads(body.decode("utf-8-sig"))
    if not isinstance(payload, list) or len(payload) < 2 or not isinstance(payload[1], list):
        raise ValueError("Unexpected World Bank country metadata response.")
    records: dict[str, dict[str, str]] = {}
    for row in payload[1]:
        code = str(row.get("id", "")).strip().upper()
        region = str((row.get("region") or {}).get("value", "")).strip()
        if not re.fullmatch(r"[A-Z]{3}", code) or not region or region == "Aggregates":
            continue
        if code in records:
            raise ValueError(f"Duplicate World Bank country code: {code}")
        records[code] = {
            "display_name": str(row.get("name", "")).strip(),
            "region": region,
            "iso2": str(row.get("iso2Code", "")).strip().upper(),
        }
    return records


def parse_migrant_destination_stock(body: bytes) -> dict[str, dict[str, object]]:
    workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    if "Table 1" not in workbook.sheetnames:
        raise ValueError("UN migrant-stock workbook does not contain Table 1.")
    rows = list(workbook["Table 1"].iter_rows(values_only=True))
    header_index = next(
        (index for index, row in enumerate(rows[:30]) if "Location code" in row and 2024 in row),
        None,
    )
    if header_index is None:
        raise ValueError("UN migrant-stock Table 1 headers changed.")
    header = list(rows[header_index])
    name_column = header.index("Region, development group, country or area")
    m49_column = header.index("Location code")
    value_column = header.index(2024)
    records: dict[str, dict[str, object]] = {}
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if row[m49_column] is None or row[value_column] is None:
            continue
        try:
            m49 = str(int(row[m49_column])).zfill(3)
            value = float(row[value_column])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(value) or value < 0:
            continue
        records[m49] = {
            "name": str(row[name_column]).strip(),
            "migrant_stock": value,
            "reference_year": 2024,
            "record_locator": f"Table 1!R{row_number}C{value_column + 1}",
        }
    return records


def build_country_registry(
    m49_body: bytes,
    world_bank_body: bytes,
    *,
    policy: dict[str, object],
) -> dict[str, CountryRegistryRecord]:
    m49 = parse_m49_registry(m49_body)
    world_bank = parse_world_bank_country_metadata(world_bank_body)
    excluded = {
        str(item["code"]): str(item["reason"])
        for item in policy.get("excluded_entities", [])  # type: ignore[union-attr]
    }
    aliases_by_code: dict[str, list[str]] = {}
    for alias, code in dict(policy.get("aliases", {})).items():  # type: ignore[arg-type]
        aliases_by_code.setdefault(str(code), []).append(str(alias))
    records: dict[str, CountryRegistryRecord] = {}
    for code, un_record in sorted(m49.items()):
        wb_record = world_bank.get(code)
        if not wb_record:
            continue
        entity_type = "excluded_entity" if code in excluded else "country"
        records[code] = CountryRegistryRecord(
            code=code,
            display_name=wb_record["display_name"] or un_record["display_name"],
            un_m49=un_record["un_m49"],
            region=wb_record["region"],
            subregion=un_record["subregion"],
            entity_type=entity_type,
            active=True,
            source_codes={
                "world_bank": code,
                "un_migrant_stock": un_record["un_m49"],
                "un_m49": un_record["un_m49"],
            },
            aliases=tuple(sorted(set(aliases_by_code.get(code, [])))),
        )
    validate_country_registry(records, excluded)
    return records


def validate_country_registry(
    records: dict[str, CountryRegistryRecord], excluded: dict[str, str]
) -> None:
    if len(records) != len(set(records)):
        raise ValueError("Duplicate ISO3 country codes.")
    source_mappings: dict[tuple[str, str], str] = {}
    for code, record in records.items():
        if not re.fullmatch(r"[A-Z]{3}", code) or not record.display_name:
            raise ValueError(f"Invalid country registry record: {code}")
        if not record.region:
            raise ValueError(f"Missing region for {code}")
        for source_id, source_code in record.source_codes.items():
            key = source_id, source_code
            if key in source_mappings:
                raise ValueError(f"Duplicate source mapping {source_id}:{source_code}")
            source_mappings[key] = code
        if record.entity_type == "excluded_entity" and code not in excluded:
            raise ValueError(f"Unexplained non-country entity: {code}")


def select_candidates(
    registry: dict[str, CountryRegistryRecord],
    migrant_body: bytes,
    policy: dict[str, object],
    *,
    candidate_limit: int | None = None,
) -> list[dict[str, object]]:
    migrant_by_m49 = parse_migrant_destination_stock(migrant_body)
    target = candidate_limit or int(policy["candidate_pool_target"])
    existing_codes = {str(code) for code in policy.get("existing_country_codes", [])}
    ranked = []
    for code, record in registry.items():
        if record.entity_type != "country":
            continue
        migrant = migrant_by_m49.get(record.un_m49)
        if not migrant:
            continue
        ranked.append((float(migrant["migrant_stock"]), code, record, migrant))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    migration_ranks = {code: rank for rank, (_, code, _, _) in enumerate(ranked, start=1)}
    unresolved_existing = existing_codes - set(registry)
    if unresolved_existing:
        raise ValueError(
            f"Active-release countries missing from canonical registry: {sorted(unresolved_existing)}"
        )
    invalid_existing = {code for code in existing_codes if registry[code].entity_type != "country"}
    if invalid_existing:
        raise ValueError(
            f"Active-release countries excluded by entity policy: {sorted(invalid_existing)}"
        )
    selected_codes = set(existing_codes)
    for _, code, _, _ in ranked:
        if len(selected_codes) >= target:
            break
        selected_codes.add(code)
    ordered = [item for item in ranked if item[1] in selected_codes]
    if len(ordered) < target:
        raise ValueError(f"Only {len(ordered)} eligible migrant destinations; need {target}.")
    candidates = []
    for inclusion_rank, (stock, code, record, migrant) in enumerate(ordered, start=1):
        existing = code in existing_codes
        reason_codes = ["official_migrant_stock_rank"]
        if existing:
            reason_codes.append("existing_konsider_country")
        candidates.append(
            {
                "code": code,
                "display_name": record.display_name,
                "region": record.region,
                "subregion": record.subregion,
                "inclusion_rank": inclusion_rank,
                "signals": {
                    "international_migrant_stock_rank": migration_ranks[code],
                    "international_migrant_stock_2024": int(stock),
                    "existing_konsider_country": existing,
                },
                "inclusion_reason_codes": reason_codes,
                "source_record": migrant["record_locator"],
            }
        )
    if not existing_codes <= {str(item["code"]) for item in candidates}:
        missing = existing_codes - {str(item["code"]) for item in candidates}
        raise ValueError(f"Candidate limit displaced active-release countries: {sorted(missing)}")
    return candidates


def _world_bank_rows(body: bytes) -> list[dict[str, object]]:
    if body.startswith(b"PK"):
        return _world_bank_zip_rows(body)
    payload = json.loads(body.decode("utf-8-sig"))
    if not isinstance(payload, list) or len(payload) < 2 or not isinstance(payload[1], list):
        raise ValueError("Unexpected World Bank indicator response.")
    return payload[1]


def _world_bank_zip_rows(body: bytes) -> list[dict[str, object]]:
    with zipfile.ZipFile(io.BytesIO(body)) as archive:
        candidates = [
            name
            for name in archive.namelist()
            if name.lower().endswith(".csv") and not Path(name).name.lower().startswith("metadata_")
        ]
        if len(candidates) != 1:
            raise ValueError(
                f"Expected one World Bank indicator CSV in ZIP; found {len(candidates)}."
            )
        text = archive.read(candidates[0]).decode("utf-8-sig")
    lines = text.splitlines()
    header_index = next(
        (
            index
            for index, line in enumerate(lines[:20])
            if line.startswith('"Country Name","Country Code","Indicator Name","Indicator Code"')
        ),
        None,
    )
    if header_index is None:
        raise ValueError("World Bank bulk CSV headers changed.")
    output = []
    for row in csv.DictReader(lines[header_index:]):
        code = str(row.get("Country Code", "")).strip().upper()
        indicator = str(row.get("Indicator Code", "")).strip()
        if not re.fullmatch(r"[A-Z]{3}", code):
            continue
        for year, raw_value in row.items():
            if not year or not year.isdigit() or not raw_value:
                continue
            try:
                value = float(raw_value)
            except ValueError:
                continue
            output.append(
                {
                    "countryiso3code": code,
                    "date": year,
                    "value": value,
                    "indicator": {"id": indicator},
                }
            )
    return output


def _wdi_series(
    bodies: bytes | Iterable[bytes], candidate_codes: set[str]
) -> dict[str, list[dict[str, object]]]:
    result: dict[str, list[dict[str, object]]] = {code: [] for code in candidate_codes}
    body_items = [bodies] if isinstance(bodies, bytes) else bodies
    for body in body_items:
        for row in _world_bank_rows(body):
            code = str(row.get("countryiso3code", "")).upper()
            if code in result:
                result[code].append(row)
    return result


def _latest_non_null(rows: Iterable[dict[str, object]]) -> tuple[float, int] | None:
    valid = []
    for row in rows:
        if row.get("value") is None:
            continue
        try:
            value = float(row["value"])
            year = int(str(row["date"]))
        except (TypeError, ValueError):
            continue
        if math.isfinite(value):
            valid.append((year, value))
    if not valid:
        return None
    year, value = max(valid)
    return value, year


def _criterion_result(
    *,
    criterion_id: str,
    source_id: str,
    raw_record_present: bool,
    raw_value_non_null: bool,
    value: float | None,
    reference_years: list[int],
    as_of_year: int,
    parse_error: str | None = None,
) -> dict[str, object]:
    finite = value is not None and math.isfinite(value)
    value_range = RANGES[criterion_id]
    range_valid = finite and value_range[0] <= value <= value_range[1]  # type: ignore[operator]
    freshness_limit = FRESHNESS_MAX_AGE[criterion_id]
    fresh = (
        bool(reference_years)
        and max(as_of_year - year for year in reference_years) <= freshness_limit
    )
    parser_succeeded = parse_error is None and raw_value_non_null and finite
    validation_succeeded = parser_succeeded and range_valid and fresh
    score_produced = False
    score_error = None
    if validation_succeeded:
        try:
            direction = (
                "lower_is_better"
                if criterion_id
                in {
                    "ambient_pm25_population_weighted",
                    "intentional_homicide_rate",
                    "household_consumption_price_level_us_100",
                }
                else "higher_is_better"
            )
            _, anchors = CURRENT_THRESHOLD_METHODS[criterion_id]
            _algorithm_scores(
                [float(value)],
                "threshold",
                direction,
                anchors,
                metric_id=criterion_id,
                broad_icp=criterion_id == "household_consumption_price_level_us_100",
            )
            score_produced = True
        except Exception as exc:  # pragma: no cover - defensive reporting
            score_error = f"{type(exc).__name__}: {exc}"
    if parse_error:
        status = "parse_failed"
    elif not raw_record_present:
        status = "missing"
    elif not raw_value_non_null:
        status = "invalid_value"
    elif not finite or not range_valid:
        status = "invalid_value"
    elif not fresh:
        status = "stale"
    elif not validation_succeeded:
        status = "validation_failed"
    elif not score_produced:
        status = "score_failed"
    else:
        status = "available"
    reasons = [] if status == "available" else [status]
    return {
        "criterion_id": criterion_id,
        "source_id": source_id,
        "source_country_identifier_resolved": True,
        "raw_source_record_present": raw_record_present,
        "raw_value_non_null": raw_value_non_null,
        "value_finite": finite,
        "unit_accepted": True,
        "reference_period_present": bool(reference_years),
        "reference_years": sorted(reference_years),
        "freshness_policy_passed": fresh,
        "parser_succeeded": parser_succeeded,
        "validation_succeeded": validation_succeeded,
        "score_produced": score_produced,
        "criterion_ready": True,
        "status": status,
        "exclusion_reason_codes": reasons,
        "error": parse_error or score_error,
    }


def _wdi_single_criterion(
    criterion_id: str,
    source_id: str,
    bodies: bytes | Iterable[bytes],
    candidate_codes: set[str],
    as_of_year: int,
) -> dict[str, dict[str, object]]:
    try:
        series = _wdi_series(bodies, candidate_codes)
    except Exception as exc:
        error = f"{type(exc).__name__}: {exc}"
        return {
            code: _criterion_result(
                criterion_id=criterion_id,
                source_id=source_id,
                raw_record_present=False,
                raw_value_non_null=False,
                value=None,
                reference_years=[],
                as_of_year=as_of_year,
                parse_error=error,
            )
            for code in candidate_codes
        }
    output = {}
    for code, rows in series.items():
        latest = _latest_non_null(rows)
        output[code] = _criterion_result(
            criterion_id=criterion_id,
            source_id=source_id,
            raw_record_present=bool(rows),
            raw_value_non_null=latest is not None,
            value=latest[0] if latest else None,
            reference_years=[latest[1]] if latest else [],
            as_of_year=as_of_year,
        )
    return output


def _icp_criterion(
    bodies: list[bytes], candidate_codes: set[str], as_of_year: int
) -> dict[str, dict[str, object]]:
    try:
        ppp_bodies, exchange_bodies = _body_groups(bodies, 2)
        ppp = _wdi_series(ppp_bodies, candidate_codes)
        exchange = _wdi_series(exchange_bodies, candidate_codes)
    except Exception as exc:
        error = f"{type(exc).__name__}: {exc}"
        return {
            code: _criterion_result(
                criterion_id="household_consumption_price_level_us_100",
                source_id="world_bank_icp",
                raw_record_present=False,
                raw_value_non_null=False,
                value=None,
                reference_years=[],
                as_of_year=as_of_year,
                parse_error=error,
            )
            for code in candidate_codes
        }
    output = {}
    for code in candidate_codes:
        ppp_values = {
            int(str(row["date"])): float(row["value"])
            for row in ppp[code]
            if row.get("value") is not None
        }
        exchange_values = {
            int(str(row["date"])): float(row["value"])
            for row in exchange[code]
            if row.get("value") is not None
        }
        years = sorted(set(ppp_values) & set(exchange_values))
        value = None
        if years:
            year = years[-1]
            exchange_value = exchange_values[year]
            value = ppp_values[year] / exchange_value * 100 if exchange_value else math.nan
        output[code] = _criterion_result(
            criterion_id="household_consumption_price_level_us_100",
            source_id="world_bank_icp",
            raw_record_present=bool(ppp[code]) and bool(exchange[code]),
            raw_value_non_null=bool(years),
            value=value,
            reference_years=[years[-1]] if years else [],
            as_of_year=as_of_year,
        )
    return output


def _body_groups(bodies: list[bytes], group_count: int) -> list[list[bytes]]:
    if len(bodies) % group_count:
        raise ValueError(
            f"Cannot divide {len(bodies)} source artifacts into {group_count} indicator groups."
        )
    size = len(bodies) // group_count
    return [bodies[index * size : (index + 1) * size] for index in range(group_count)]


def _wbl_criterion(
    body: bytes, candidate_codes: set[str], as_of_year: int
) -> dict[str, dict[str, object]]:
    rows_by_code: dict[str, list[tuple[int, float]]] = {code: [] for code in candidate_codes}
    raw_codes: set[str] = set()
    try:
        workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        rows = list(workbook["WBL Economy Scores"].iter_rows(values_only=True))
        header_index = next(i for i, row in enumerate(rows[:20]) if row and row[0] == "Economy")
        headers = list(rows[header_index])
        iso_column = headers.index("ISO Code")
        year_column = headers.index("Report Year")
        score_column = headers.index("I. Economy LF Index")
        for row in rows[header_index + 1 :]:
            code = str(row[iso_column]).strip().upper() if row[iso_column] else ""
            if code not in candidate_codes:
                continue
            raw_codes.add(code)
            if isinstance(row[score_column], (int, float)):
                rows_by_code[code].append((int(row[year_column]), float(row[score_column])))
    except Exception as exc:
        error = f"{type(exc).__name__}: {exc}"
        return {
            code: _criterion_result(
                criterion_id="women_legal_economic_equality",
                source_id="world_bank_wbl",
                raw_record_present=False,
                raw_value_non_null=False,
                value=None,
                reference_years=[],
                as_of_year=as_of_year,
                parse_error=error,
            )
            for code in candidate_codes
        }
    output = {}
    for code in candidate_codes:
        latest = max(rows_by_code[code], default=None)
        output[code] = _criterion_result(
            criterion_id="women_legal_economic_equality",
            source_id="world_bank_wbl",
            raw_record_present=code in raw_codes,
            raw_value_non_null=latest is not None,
            value=latest[1] if latest else None,
            reference_years=[2025] if latest else [],
            as_of_year=as_of_year,
        )
    return output


def _infrastructure_criterion(
    bodies: list[bytes], candidate_codes: set[str], as_of_year: int
) -> dict[str, dict[str, object]]:
    try:
        components = [_wdi_series(group, candidate_codes) for group in _body_groups(bodies, 3)]
    except Exception as exc:
        error = f"{type(exc).__name__}: {exc}"
        return {
            code: _criterion_result(
                criterion_id="infrastructure_readiness_composite",
                source_id="world_bank_infrastructure",
                raw_record_present=False,
                raw_value_non_null=False,
                value=None,
                reference_years=[],
                as_of_year=as_of_year,
                parse_error=error,
            )
            for code in candidate_codes
        }
    output = {}
    for code in candidate_codes:
        latest = [_latest_non_null(component[code]) for component in components]
        value = None
        years = []
        if all(item is not None for item in latest):
            values = [item[0] for item in latest if item]
            years = [item[1] for item in latest if item]
            internet = min(max(values[0], 0), 100)
            broadband = _piecewise(values[1], ((0, 0), (10, 25), (20, 50), (30, 75), (40, 100)))
            lpi = _piecewise(values[2], ((1, 0), (2, 25), (3, 50), (4, 75), (5, 100)))
            value = (internet + broadband + lpi) / 3
        output[code] = _criterion_result(
            criterion_id="infrastructure_readiness_composite",
            source_id="world_bank_infrastructure",
            raw_record_present=all(bool(component[code]) for component in components),
            raw_value_non_null=all(item is not None for item in latest),
            value=value,
            reference_years=years,
            as_of_year=as_of_year,
        )
    return output


def _piecewise(value: float, anchors: tuple[tuple[float, float], ...]) -> float:
    if value <= anchors[0][0]:
        return anchors[0][1]
    if value >= anchors[-1][0]:
        return anchors[-1][1]
    for (x0, y0), (x1, y1) in zip(anchors, anchors[1:], strict=False):
        if x0 <= value <= x1:
            return y0 + (value - x0) / (x1 - x0) * (y1 - y0)
    raise AssertionError("piecewise interpolation failed")


def _artifact_bodies(
    artifacts: list[RawArtifact],
    raw_repository: RawArtifactRepository,
    registrations: dict[str, SourceRegistration],
) -> dict[str, list[bytes]]:
    grouped: dict[str, list[RawArtifact]] = {}
    for artifact in artifacts:
        grouped.setdefault(artifact.source_id, []).append(artifact)
    result = {}
    for source_id, registration in registrations.items():
        order = {url: index for index, url in enumerate(registration.download_urls)}
        ordered = sorted(grouped.get(source_id, []), key=lambda item: order[item.requested_url])
        if len(ordered) != len(registration.download_urls):
            raise ValueError(
                f"Expected {len(registration.download_urls)} artifacts for {source_id}; "
                f"found {len(ordered)}."
            )
        result[source_id] = [raw_repository.load(item) for item in ordered]
    return result


def _write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, rows: Iterable[dict[str, object]]) -> None:
    values = [json.dumps(row, sort_keys=True, separators=(",", ":")) for row in rows]
    path.write_text("\n".join(values) + ("\n" if values else ""), encoding="utf-8")


def _read_json(path: Path) -> object:
    return json.loads(path.read_text(encoding="utf-8"))


def _active_pointer_bytes(release_root: Path) -> bytes | None:
    path = release_root / "active.json"
    return path.read_bytes() if path.exists() else None


def audit_coverage(
    universe_path: Path | str,
    audit_id: str,
    *,
    mode: str,
    output_root: Path | str = "data/reports/country-coverage",
    raw_root: Path | str = "data/raw",
    release_root: Path | str = "data/releases",
    artifact_manifest: Path | str | None = None,
    candidate_limit: int | None = None,
    fetcher: Callable | None = None,
    clock: Callable[[], datetime] | None = None,
) -> tuple[Path, dict[str, object]]:
    """Run an online or offline coverage audit without writing a release."""

    if mode not in {"online", "offline"}:
        raise ValueError("Audit mode must be 'online' or 'offline'.")
    policy = _read_json(Path(universe_path))
    if not isinstance(policy, dict):
        raise ValueError("Country universe policy must be a JSON object.")
    enabled_criteria = enabled_criteria_from_catalog()
    raw_repository = RawArtifactRepository(raw_root)
    now = (clock or (lambda: datetime.now(UTC)))()
    active_before = _active_pointer_bytes(Path(release_root))
    artifacts: list[RawArtifact] = []
    if mode == "online":
        if fetcher is None:
            from konsider.ingestion.worker import fetch_url

            fetcher = fetch_url
        for registration in UNIVERSE_SOURCES.values():
            for url in registration.download_urls:
                result = fetcher(url)
                if len(result) == 3:
                    body, final_url, media_type = result
                    metadata = {"http_status": 200}
                else:
                    body, final_url, media_type, metadata = result
                artifacts.append(
                    raw_repository.capture(
                        registration,
                        body,
                        requested_url=url,
                        final_url=final_url,
                        retrieved_at=now.isoformat(),
                        media_type=media_type,
                        **metadata,
                    )
                )
    else:
        if artifact_manifest is None:
            raise ValueError("Offline audit requires --artifacts.")
        items = _read_json(Path(artifact_manifest))
        if not isinstance(items, list):
            raise ValueError("Artifact manifest must be a JSON array.")
        artifacts = [RawArtifact(**item) for item in items]
        if artifacts:
            now = datetime.fromisoformat(max(item.retrieved_at for item in artifacts))
    universe_bodies = _artifact_bodies(artifacts, raw_repository, UNIVERSE_SOURCES)
    registry = build_country_registry(
        universe_bodies["un_m49_country_registry"][0],
        universe_bodies["world_bank_country_metadata"][0],
        policy=policy,
    )
    candidates = select_candidates(
        registry,
        universe_bodies["un_desa_migrant_stock"][0],
        policy,
        candidate_limit=candidate_limit,
    )
    candidate_codes = {str(item["code"]) for item in candidates}
    registrations = audit_source_registrations(candidate_codes, enabled_criteria)
    if mode == "online":
        criterion_registrations = {
            source_id: registration
            for source_id, registration in registrations.items()
            if source_id not in UNIVERSE_SOURCES
        }
        for registration in criterion_registrations.values():
            for url in registration.download_urls:
                result = fetcher(url)  # type: ignore[misc]
                if len(result) == 3:
                    body, final_url, media_type = result
                    metadata = {"http_status": 200}
                else:
                    body, final_url, media_type, metadata = result
                artifacts.append(
                    raw_repository.capture(
                        registration,
                        body,
                        requested_url=url,
                        final_url=final_url,
                        retrieved_at=now.isoformat(),
                        media_type=media_type,
                        **metadata,
                    )
                )
    bodies = _artifact_bodies(artifacts, raw_repository, registrations)
    all_criterion_results = {
        "ambient_pm25_population_weighted": _wdi_single_criterion(
            "ambient_pm25_population_weighted",
            "world_bank_pm25",
            bodies["world_bank_pm25"],
            candidate_codes,
            now.year,
        ),
        "intentional_homicide_rate": _wdi_single_criterion(
            "intentional_homicide_rate",
            "unodc_homicide",
            bodies["unodc_homicide"],
            candidate_codes,
            now.year,
        ),
        "household_consumption_price_level_us_100": _icp_criterion(
            bodies["world_bank_icp"], candidate_codes, now.year
        ),
        "women_legal_economic_equality": _wbl_criterion(
            bodies["world_bank_wbl"][0], candidate_codes, now.year
        ),
        "infrastructure_readiness_composite": _infrastructure_criterion(
            bodies["world_bank_infrastructure"], candidate_codes, now.year
        ),
    }
    criterion_results = {
        criterion: all_criterion_results[criterion] for criterion in enabled_criteria
    }
    country_rows = []
    exclusions = []
    candidate_by_code = {str(item["code"]): item for item in candidates}
    for code in sorted(candidate_codes):
        results = [criterion_results[criterion][code] for criterion in enabled_criteria]
        publishable = all(item["status"] == "available" for item in results)
        reason_codes = sorted(
            {
                f"{item['criterion_id']}:{reason}"
                for item in results
                for reason in item["exclusion_reason_codes"]  # type: ignore[union-attr]
            }
        )
        row = {
            "code": code,
            "display_name": candidate_by_code[code]["display_name"],
            "region": candidate_by_code[code]["region"],
            "publishable": publishable,
            "criteria": {str(item["criterion_id"]): item for item in results},
            "exclusion_reason_codes": reason_codes,
        }
        country_rows.append(row)
        if not publishable:
            exclusions.append(
                {
                    "code": code,
                    "display_name": row["display_name"],
                    "region": row["region"],
                    "reason_codes": reason_codes,
                }
            )
    complete_count = sum(bool(item["publishable"]) for item in country_rows)
    minimum = int(policy["minimum_publishable_country_count"])
    coverage_rows = {}
    for criterion in enabled_criteria:
        values = [criterion_results[criterion][code] for code in candidate_codes]
        coverage_rows[criterion] = {
            "candidate": len(values),
            "found": sum(bool(item["raw_source_record_present"]) for item in values),
            "fresh": sum(bool(item["freshness_policy_passed"]) for item in values),
            "parsed": sum(bool(item["parser_succeeded"]) for item in values),
            "validated": sum(bool(item["validation_succeeded"]) for item in values),
            "scored": sum(bool(item["score_produced"]) for item in values),
            "available": sum(item["status"] == "available" for item in values),
            "missing": sum(item["status"] != "available" for item in values),
            "status_counts": dict(sorted(Counter(str(item["status"]) for item in values).items())),
            "otherwise_valid_excluded_only_by_this_criterion": sum(
                row["criteria"][criterion]["status"] != "available"
                and all(
                    item["status"] == "available"
                    for other, item in row["criteria"].items()
                    if other != criterion
                )
                for row in country_rows
            ),
        }
    region_candidates = Counter(str(item["region"]) for item in candidates)
    region_publishable = Counter(
        str(item["region"]) for item in country_rows if item["publishable"]
    )
    summary = {
        "schema_version": AUDIT_SCHEMA_VERSION,
        "audit_id": audit_id,
        "audit_timestamp": now.isoformat(),
        "mode": mode,
        "universe_id": policy["universe_id"],
        "selection_policy_version": policy["selection_policy_version"],
        "candidate_country_count": len(candidates),
        "enabled_criteria": enabled_criteria,
        "complete_publishable_country_count": complete_count,
        "minimum_required_country_count": minimum,
        "status": "PASS" if complete_count >= minimum else "FAIL",
        "complete_case_rule": True,
        "candidate_counts_by_region": dict(sorted(region_candidates.items())),
        "publishable_counts_by_region": dict(sorted(region_publishable.items())),
        "criterion_coverage": coverage_rows,
        "excluded_country_count": len(exclusions),
        "exclusion_reason_counts": dict(
            sorted(
                Counter(reason for item in exclusions for reason in item["reason_codes"]).items()
            )
        ),
        "source_versions": {
            source_id: registration.source_version
            for source_id, registration in sorted(registrations.items())
        },
        "raw_artifact_checksums": {
            artifact.source_id + ":" + artifact.requested_url: f"sha256:{artifact.sha256}"
            for artifact in artifacts
        },
    }
    output = Path(output_root) / audit_id
    output.mkdir(parents=True, exist_ok=False)
    _write_json(output / "summary.json", summary)
    _write_jsonl(output / "candidate-countries.jsonl", candidates)
    _write_json(output / "country-registry.json", [item.to_dict() for item in registry.values()])
    _write_json(output / "criterion-coverage.json", coverage_rows)
    _write_jsonl(output / "country-coverage.jsonl", country_rows)
    _write_jsonl(output / "exclusions.jsonl", exclusions)
    _write_json(output / "raw-artifacts.json", [item.to_dict() for item in artifacts])
    _write_json(
        output / "sources.json",
        [registration.to_dict() for registration in registrations.values()],
    )
    (output / "report.md").write_text(_render_report(summary), encoding="utf-8")
    if _active_pointer_bytes(Path(release_root)) != active_before:
        raise RuntimeError("Coverage audit changed the active release pointer.")
    return output, summary


def _render_report(summary: dict[str, object]) -> str:
    lines = [
        f"# Country coverage audit: {summary['audit_id']}",
        "",
        f"- Universe: `{summary['universe_id']}`",
        f"- Candidate countries: {summary['candidate_country_count']}",
        f"- Enabled criteria: {len(summary['enabled_criteria'])}",
        f"- Complete publishable countries: {summary['complete_publishable_country_count']}",
        f"- Minimum required: {summary['minimum_required_country_count']}",
        f"- Status: **{summary['status']}**",
        "",
        "## Criterion coverage",
        "",
        "| Criterion | Candidate | Found | Fresh | Parsed | Validated | Scored | Missing |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    coverage = summary["criterion_coverage"]
    for criterion, row in coverage.items():  # type: ignore[union-attr]
        lines.append(
            f"| `{criterion}` | {row['candidate']} | {row['found']} | {row['fresh']} | "
            f"{row['parsed']} | {row['validated']} | {row['scored']} | {row['missing']} |"
        )
    lines.extend(
        [
            "",
            "Popularity is used only to select candidates. It does not contribute to affinity scores.",
            "No missing values are imputed and no partial-country scoring is used.",
            "",
        ]
    )
    return "\n".join(lines)
