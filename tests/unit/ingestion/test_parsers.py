import io
import json
from unittest import TestCase

from openpyxl import Workbook

from konsider.ingestion.models import RawArtifact
from konsider.ingestion.parsers import (
    parse_who_air_quality,
    parse_world_bank_icp,
    parse_world_bank_migrant_stock,
    parse_world_bank_wbl,
    parse_world_bank_wgi_political_stability,
    parse_wps_index,
)


def artifact(source_id="test", index=0):
    return RawArtifact(
        f"sha256:{index}",
        source_id,
        "url",
        "url",
        "2026-01-01T00:00:00Z",
        "application/json",
        1,
        str(index),
        "v1",
        "p1",
        "unused",
    )


class ParserTests(TestCase):
    def test_who_air_selects_latest_all_area_country_value(self):
        body = json.dumps(
            {
                "value": [
                    {
                        "SpatialDimType": "COUNTRY",
                        "SpatialDim": "IND",
                        "Dim1": "RESIDENCEAREATYPE_ALL",
                        "TimeDim": 2020,
                        "NumericValue": 50,
                        "Low": 45,
                        "High": 55,
                    },
                    {
                        "SpatialDimType": "COUNTRY",
                        "SpatialDim": "IND",
                        "Dim1": "RESIDENCEAREATYPE_ALL",
                        "TimeDim": 2021,
                        "NumericValue": 48,
                        "Low": 44,
                        "High": 52,
                    },
                    {
                        "SpatialDimType": "COUNTRY",
                        "SpatialDim": "IND",
                        "Dim1": "RESIDENCEAREATYPE_RUR",
                        "TimeDim": 2022,
                        "NumericValue": 60,
                    },
                ]
            }
        ).encode()
        observations = parse_who_air_quality([artifact("who_air_quality")], [body])
        self.assertEqual(
            [(item.country_code, item.value, item.reference_end) for item in observations],
            [("IND", 48.0, "2021-12-31")],
        )
        self.assertEqual(observations[0].observation_type, "modelled")
        self.assertEqual(observations[0].raw_artifact_ids, ("sha256:0",))
        self.assertEqual(observations[0].source_records[0].locator, "$.value[1]")

    def test_icp_preserves_both_inputs_and_derives_price_level(self):
        ppp = json.dumps([{}, [{"countryiso3code": "IND", "date": "2021", "value": 20}]]).encode()
        exchange = json.dumps(
            [{}, [{"countryiso3code": "IND", "date": "2021", "value": 80}]]
        ).encode()
        observations = parse_world_bank_icp(
            [artifact("world_bank_icp", 1), artifact("world_bank_icp", 2)], [ppp, exchange]
        )
        self.assertEqual(observations[0].value, 25.0)
        self.assertEqual(len(observations[0].raw_artifact_ids), 2)
        self.assertEqual(len(observations[0].source_records), 2)

    def test_wps_finds_named_columns_without_sheet_position_assumptions(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["notes"])
        sheet.append(["Country", "WPS Index Score"])
        sheet.append(["India", 0.61])
        output = io.BytesIO()
        workbook.save(output)
        observations = parse_wps_index([artifact("wps_index")], [output.getvalue()])
        self.assertEqual((observations[0].country_code, observations[0].value), ("IND", 0.61))

    def test_wbl_selects_latest_report_row_per_country(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "WBL Economy Scores"
        sheet.append(["notes"])
        sheet.append(["more notes"])
        sheet.append(["methodology"])
        sheet.append(["Economy", "ISO Code", "Report Year", "I. Economy LF Index"])
        sheet.append(["India", "IND", 2025, 72.5])
        sheet.append(["India", "IND", 2026, 75.0])
        output = io.BytesIO()
        workbook.save(output)

        observations = parse_world_bank_wbl([artifact("world_bank_wbl")], [output.getvalue()])

        self.assertEqual(len(observations), 1)
        self.assertEqual((observations[0].country_code, observations[0].value), ("IND", 75.0))
        self.assertEqual(
            observations[0].source_records[0].record_id, "IND|WBL_LF_INDEX|report-2026"
        )

    def test_wgi_retains_published_uncertainty(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "pv"
        sheet.append(
            [
                "Economy (code)",
                "Year",
                "Governance estimate (approx. -2.5 to +2.5)",
                "Standard error (estimate)",
                "Lower threshold (90% conf. int. estimate)",
                "Upper threshold (90% conf. int. estimate)",
            ]
        )
        sheet.append(["IND", 2024, 0.25, 0.1, 0.08, 0.42])
        output = io.BytesIO()
        workbook.save(output)

        observations = parse_world_bank_wgi_political_stability(
            [artifact("world_bank_wgi_political_stability")], [output.getvalue()]
        )

        self.assertEqual((observations[0].country_code, observations[0].value), ("IND", 0.25))
        self.assertEqual(
            (observations[0].lower_bound, observations[0].upper_bound),
            (0.08, 0.42),
        )
        self.assertEqual(observations[0].components[0].value, 0.1)

    def test_migrant_stock_selects_latest_value(self):
        body = json.dumps(
            [
                {},
                [
                    {"countryiso3code": "IND", "date": "2023", "value": 1.4},
                    {"countryiso3code": "IND", "date": "2024", "value": 1.6},
                ],
            ]
        ).encode()

        observations = parse_world_bank_migrant_stock(
            [artifact("world_bank_migrant_stock")], [body]
        )

        self.assertEqual((observations[0].country_code, observations[0].value), ("IND", 1.6))
        self.assertIn("preference_property_not_universal_quality", observations[0].quality_flags)
