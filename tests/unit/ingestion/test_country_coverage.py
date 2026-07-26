import io
import json
import zipfile
from datetime import UTC, datetime

from openpyxl import Workbook

from konsider.ingestion.country_coverage import (
    _wdi_single_criterion,
    _world_bank_rows,
    audit_coverage,
    enabled_criteria_from_catalog,
    parse_m49_registry,
    select_candidates,
)


def _m49_html(records):
    rows = "".join(
        "<tr>"
        f"<td>001</td><td>World</td><td>150</td><td>Europe</td>"
        f"<td>039</td><td>Southern Europe</td><td></td><td></td>"
        f"<td>{name}</td><td>{index:03d}</td><td>{code[:2]}</td><td>{code}</td>"
        "<td></td><td></td><td></td>"
        "</tr>"
        for index, code, name in records
    )
    return (
        '<table id="downloadTableEN"><tr>'
        "<th>Global Code</th><th>Global Name</th><th>Region Code</th><th>Region Name</th>"
        "<th>Sub-region Code</th><th>Sub-region Name</th>"
        "<th>Intermediate Region Code</th><th>Intermediate Region Name</th>"
        "<th>Country or Area</th><th>M49 Code</th><th>ISO-alpha2 Code</th>"
        "<th>ISO-alpha3 Code</th><th>Least Developed Countries (LDC)</th>"
        "<th>Land Locked Developing Countries (LLDC)</th>"
        f"<th>Small Island Developing States (SIDS)</th></tr>{rows}</table>"
    ).encode()


def _migrant_workbook(records):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table 1"
    for _ in range(10):
        sheet.append([])
    sheet.append(
        [
            None,
            "Region, development group, country or area",
            "Coverage",
            "Data type",
            "Location code",
            2024,
        ]
    )
    for index, _, name in records:
        sheet.append([None, name, None, None, index, 1_000_000 - index])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _wbl_workbook(codes):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WBL Economy Scores"
    sheet.append(["notes"])
    sheet.append(["Economy", "ISO Code", "Report Year", "I. Economy LF Index"])
    for code in codes:
        sheet.append([code, code, 2026, 75])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _wgi_workbook(codes):
    workbook = Workbook()
    workbook.remove(workbook.active)
    headers = [
        "Economy (code)",
        "Year",
        "Governance estimate (approx. -2.5 to +2.5)",
    ]
    for name in ("pv", "rl"):
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        for code in codes:
            sheet.append([code, 2024, 0.5])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _wdi(codes, value=50, year=2025):
    return json.dumps(
        [
            {"page": 1, "pages": 1},
            [
                {
                    "countryiso3code": code,
                    "date": str(year),
                    "value": value,
                    "indicator": {"id": "fixture"},
                }
                for code in codes
            ],
        ]
    ).encode()


def _records(count):
    records = []
    for index in range(1, count + 1):
        first = chr(65 + (index // 676) % 26)
        second = chr(65 + (index // 26) % 26)
        third = chr(65 + index % 26)
        records.append((index, first + second + third, f"Country {index:03d}"))
    return records


def test_m49_registry_rejects_unexpectedly_small_table():
    records = _records(199)
    try:
        parse_m49_registry(_m49_html(records))
    except ValueError as exc:
        assert "unexpectedly contained" in str(exc)
    else:
        raise AssertionError("Expected small M49 table to be rejected")


def test_enabled_criteria_are_catalog_driven(tmp_path):
    catalog = tmp_path / "catalog.json"
    catalog.write_text(
        json.dumps(
            {
                "criteria": [
                    {
                        "id": "ambient_pm25_population_weighted",
                        "ready": True,
                        "default_enabled": True,
                    },
                    {
                        "id": "uhc_service_coverage_index",
                        "ready": False,
                        "default_enabled": False,
                    },
                ]
            }
        ),
        encoding="utf-8",
    )
    assert enabled_criteria_from_catalog(catalog) == ["ambient_pm25_population_weighted"]


def test_candidate_selection_is_deterministic_and_retains_existing():
    records = _records(205)
    policy = {
        "candidate_pool_target": 150,
        "existing_country_codes": [records[180][1]],
    }
    registry = {
        code: type(
            "Record",
            (),
            {
                "entity_type": "country",
                "un_m49": f"{index:03d}",
                "display_name": name,
                "region": "Europe",
                "subregion": "Southern Europe",
            },
        )()
        for index, code, name in records
    }
    candidates = select_candidates(registry, _migrant_workbook(records), policy)
    assert len(candidates) == 150
    assert records[180][1] in {row["code"] for row in candidates}
    assert candidates == select_candidates(registry, _migrant_workbook(records), policy)


def test_online_and_offline_audits_match_and_do_not_activate_release(tmp_path, monkeypatch):
    records = _records(205)
    codes = [code for _, code, _ in records[:150]]
    universe = tmp_path / "universe.json"
    universe.write_text(
        json.dumps(
            {
                "universe_id": "test_v1",
                "selection_policy_version": "test-policy-v1",
                "minimum_publishable_country_count": 100,
                "candidate_pool_target": 150,
                "existing_country_codes": [],
                "aliases": {},
                "excluded_entities": [],
            }
        ),
        encoding="utf-8",
    )
    releases = tmp_path / "releases"
    releases.mkdir()
    active = releases / "active.json"
    active.write_text('{"release_id":"before"}\n', encoding="utf-8")
    bodies = {
        "unstats.un.org": _m49_html(records),
        "ims_stock": _migrant_workbook(records),
        "/v2/country?": json.dumps(
            [
                {},
                [
                    {
                        "id": code,
                        "iso2Code": code[:2],
                        "name": name,
                        "region": {"value": "Europe & Central Asia"},
                    }
                    for _, code, name in records
                ],
            ]
        ).encode(),
        "WBL26": _wbl_workbook(codes),
        "wgidataset_with_sourcedata": _wgi_workbook(codes),
    }

    def fetcher(url):
        for marker, body in bodies.items():
            if marker in url:
                media = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    if marker in {"ims_stock", "WBL26"}
                    else "application/json"
                )
                return body, url, media, {"http_status": 200}
        if "LP.LPI.INFR.XQ" in url:
            return _wdi(codes, 3, 2023), url, "application/json", {"http_status": 200}
        if "PA.NUS.PRVT.PP" in url:
            return _wdi(codes, 50), url, "application/json", {"http_status": 200}
        if "PA.NUS.FCRF" in url:
            return _wdi(codes, 100), url, "application/json", {"http_status": 200}
        return _wdi(codes), url, "application/json", {"http_status": 200}

    def clock():
        return datetime(2026, 7, 23, tzinfo=UTC)

    online_path, online = audit_coverage(
        universe,
        "online",
        mode="online",
        output_root=tmp_path / "reports",
        raw_root=tmp_path / "raw",
        release_root=releases,
        fetcher=fetcher,
        clock=clock,
    )
    assert online["status"] == "PASS"
    assert online["complete_publishable_country_count"] == 150
    assert active.read_text(encoding="utf-8") == '{"release_id":"before"}\n'

    bodies["WBL26"] = _wbl_workbook(codes[:99])
    _, failed = audit_coverage(
        universe,
        "below-minimum",
        mode="online",
        output_root=tmp_path / "reports",
        raw_root=tmp_path / "raw",
        release_root=releases,
        fetcher=fetcher,
        clock=clock,
    )
    assert failed["status"] == "FAIL"
    assert failed["complete_publishable_country_count"] == 99
    assert active.read_text(encoding="utf-8") == '{"release_id":"before"}\n'

    _, offline = audit_coverage(
        universe,
        "offline",
        mode="offline",
        output_root=tmp_path / "reports",
        raw_root=tmp_path / "raw",
        release_root=releases,
        artifact_manifest=online_path / "raw-artifacts.json",
    )
    ignored = {"audit_id", "mode"}
    assert {k: v for k, v in online.items() if k not in ignored} == {
        k: v for k, v in offline.items() if k not in ignored
    }
    assert active.read_text(encoding="utf-8") == '{"release_id":"before"}\n'


def test_coverage_distinguishes_stale_from_missing():
    body = json.dumps(
        [
            {},
            [
                {
                    "countryiso3code": "AAA",
                    "date": "2020",
                    "value": 50,
                    "indicator": {"id": "fixture"},
                }
            ],
        ]
    ).encode()
    results = _wdi_single_criterion(
        "ambient_pm25_population_weighted",
        "world_bank_pm25",
        body,
        {"AAA", "BBB"},
        2026,
    )
    assert results["AAA"]["status"] == "stale"
    assert results["BBB"]["status"] == "missing"


def test_coverage_records_parse_and_score_failures(monkeypatch):
    malformed = _wdi_single_criterion(
        "ambient_pm25_population_weighted",
        "world_bank_pm25",
        b"not-json",
        {"AAA"},
        2026,
    )
    assert malformed["AAA"]["status"] == "parse_failed"

    monkeypatch.setattr(
        "konsider.ingestion.country_coverage._algorithm_scores",
        lambda *args, **kwargs: (_ for _ in ()).throw(ValueError("fixture score failure")),
    )
    score_failure = _wdi_single_criterion(
        "ambient_pm25_population_weighted",
        "world_bank_pm25",
        _wdi(["AAA"], year=2025),
        {"AAA"},
        2026,
    )
    assert score_failure["AAA"]["status"] == "score_failed"


def test_world_bank_bulk_zip_is_parsed_as_indicator_rows():
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        archive.writestr(
            "API_TEST_DS2_en_csv_v2.csv",
            '"Data Source","World Development Indicators"\n'
            '"Last Updated Date","2026-07-23"\n\n'
            '"Country Name","Country Code","Indicator Name","Indicator Code","2024","2025"\n'
            '"Alpha","AAA","Fixture","TEST.SERIES","10","11"\n',
        )
        archive.writestr("Metadata_Country_API_TEST.csv", "metadata")
    rows = _world_bank_rows(output.getvalue())
    assert [(row["countryiso3code"], row["date"], row["value"]) for row in rows] == [
        ("AAA", "2024", 10.0),
        ("AAA", "2025", 11.0),
    ]
